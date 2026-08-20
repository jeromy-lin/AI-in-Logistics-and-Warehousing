# ============================================================
# 物流業 WMS AI 智慧成本分析實作
# XGBoost - Cost-to-Serve Prediction
# Public Teaching Version | Warm Dashboard Style
# 作者：國立雲林科技大學電機系 林家仁
# ★ 與 Multiple Regression / Random Forest 使用相同資料
# ★ 相同 X / y
# ★ 相同 Customer_ID Group Split
# ★ 相同 R² / MAE / RMSE / MAPE
# ============================================================


# ============================================================
# STEP 0 - 安裝 / 載入套件
# ============================================================

!pip -q install xgboost openpyxl

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import warnings

warnings.filterwarnings("ignore")

from IPython.display import display, HTML
from google.colab import files

from sklearn.model_selection import GroupShuffleSplit
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline

from sklearn.metrics import (
    r2_score,
    mean_absolute_error,
    mean_squared_error,
    mean_absolute_percentage_error
)

from xgboost import XGBRegressor

from matplotlib.colors import LinearSegmentedColormap


# ============================================================
# STEP 1 - 視覺設計
# ============================================================

COLORS = {

    "cream": "#FFF8F0",
    "cream2": "#FFF2E5",

    "peach": "#F7C9A9",
    "peach_dark": "#E9A77F",

    "coral": "#E98F7C",
    "terracotta": "#C96F5B",

    "rose": "#D7A3A9",

    "gold": "#D9A15B",

    "sage": "#A8B89A",
    "sage_dark": "#7F9273",

    "brown": "#6D5142",
    "brown_dark": "#4F3A30",

    "gray": "#746C67",

    "white": "#FFFFFF"
}


SIZE_COLORS = {

    "S60": "#A8B89A",

    "S90": "#F2C879",

    "S120": "#E9A77F",

    "S150": "#D77F73"
}


REGION_EN = {

    "北部": "North",

    "中部": "Central",

    "南部": "South",

    "東部": "East",

    "離島": "Offshore"
}


TEMPERATURE_EN = {

    "常溫": "Ambient",

    "低溫": "Chilled"
}


INDUSTRY_EN = {

    "生醫用品": "Biomedical",

    "食品冷鏈": "Cold Chain Food",

    "書籍文具": "Books & Stationery",

    "汽車零件": "Auto Parts",

    "電商零售": "E-commerce",

    "美妝保健": "Beauty & Health",

    "服飾鞋包": "Fashion",

    "3C家電": "Electronics"
}


REGION_COLORS = {

    "North": "#A8B89A",

    "Central": "#F2C879",

    "South": "#E9A77F",

    "East": "#D7A3A9",

    "Offshore": "#C96F5B"
}


TEMP_COLORS = {

    "Ambient": "#E9A77F",

    "Chilled": "#A8B89A"
}


warm_cmap = LinearSegmentedColormap.from_list(

    "warm_gradient",

    [
        "#FFF8F0",
        "#F7C9A9",
        "#E98F7C"
    ]
)


plt.rcParams["figure.figsize"] = (10, 6)
plt.rcParams["axes.titleweight"] = "bold"
plt.rcParams["axes.titlesize"] = 15
plt.rcParams["axes.labelsize"] = 11
plt.rcParams["axes.edgecolor"] = COLORS["brown"]
plt.rcParams["axes.labelcolor"] = COLORS["brown_dark"]
plt.rcParams["xtick.color"] = COLORS["gray"]
plt.rcParams["ytick.color"] = COLORS["gray"]
plt.rcParams["font.family"] = "DejaVu Sans"


# ============================================================
# STEP 2 - Dashboard Functions
# ============================================================

def dashboard_title(title, subtitle=""):

    html = f"""
    <div style="
        background:linear-gradient(135deg,#FFF8F0,#F7C9A9);
        border-left:8px solid #C96F5B;
        padding:22px 28px;
        margin:16px 0 22px 0;
        border-radius:14px;
        box-shadow:0 3px 10px rgba(100,70,50,0.10);
    ">

        <div style="
            font-size:26px;
            font-weight:800;
            color:#4F3A30;
            margin-bottom:6px;
        ">
            {title}
        </div>

        <div style="
            font-size:15px;
            color:#746C67;
            line-height:1.8;
        ">
            {subtitle}
        </div>

    </div>
    """

    display(HTML(html))


def section_title(number, title, description=""):

    html = f"""
    <div style="
        margin-top:26px;
        margin-bottom:14px;
    ">

        <div style="
            display:inline-block;
            background:#C96F5B;
            color:white;
            font-size:14px;
            font-weight:bold;
            padding:6px 12px;
            border-radius:18px;
            margin-bottom:8px;
        ">
            STEP {number}
        </div>

        <div style="
            font-size:21px;
            font-weight:800;
            color:#4F3A30;
        ">
            {title}
        </div>

        <div style="
            font-size:14px;
            color:#746C67;
            margin-top:5px;
            line-height:1.7;
        ">
            {description}
        </div>

    </div>
    """

    display(HTML(html))


def info_box(text):

    html = f"""
    <div style="
        background:#FFF8F0;
        border:1px solid #F0D7C4;
        padding:14px 18px;
        border-radius:10px;
        color:#6D5142;
        line-height:1.8;
        margin:10px 0 16px 0;
    ">
        💡 {text}
    </div>
    """

    display(HTML(html))


def metric_cards(metrics):

    cards = ""

    card_colors = [
        "#F7C9A9",
        "#F2C879",
        "#A8B89A",
        "#D7A3A9"
    ]

    for i, (name, value, note) in enumerate(metrics):

        color = card_colors[
            i % len(card_colors)
        ]

        cards += f"""
        <div style="
            flex:1;
            min-width:160px;
            background:white;
            border-top:6px solid {color};
            border-radius:12px;
            padding:16px;
            margin:6px;
            box-shadow:0 3px 10px rgba(80,60,50,0.08);
        ">

            <div style="
                color:#746C67;
                font-size:13px;
            ">
                {name}
            </div>

            <div style="
                color:#4F3A30;
                font-size:28px;
                font-weight:800;
                margin:6px 0;
            ">
                {value}
            </div>

            <div style="
                color:#8A7F79;
                font-size:12px;
            ">
                {note}
            </div>

        </div>
        """

    display(
        HTML(
            f"""
            <div style="
                display:flex;
                flex-wrap:wrap;
                margin:12px -6px 18px -6px;
            ">
                {cards}
            </div>
            """
        )
    )


def warm_style(
    df,
    formats=None,
    gradient_cols=None,
    hide_index=True
):

    styler = (
        df.style
        .set_table_styles([
            {
                "selector": "thead th",
                "props": [
                    ("background-color", COLORS["terracotta"]),
                    ("color", "white"),
                    ("font-weight", "bold"),
                    ("text-align", "center"),
                    ("padding", "10px")
                ]
            },
            {
                "selector": "tbody td",
                "props": [
                    ("padding", "8px"),
                    ("border-bottom", "1px solid #F1E4DA"),
                    ("color", COLORS["brown_dark"])
                ]
            },
            {
                "selector": "tbody tr:nth-child(even)",
                "props": [
                    ("background-color", COLORS["cream"])
                ]
            },
            {
                "selector": "tbody tr:hover",
                "props": [
                    ("background-color", COLORS["cream2"])
                ]
            }
        ])
    )

    if formats:
        styler = styler.format(formats)

    if gradient_cols:

        valid_cols = [
            c
            for c in gradient_cols
            if c in df.columns
        ]

        if valid_cols:

            styler = styler.background_gradient(
                cmap=warm_cmap,
                subset=valid_cols
            )

    if hide_index:

        try:
            styler = styler.hide(axis="index")
        except:
            pass

    return styler


# ============================================================
# STEP 3 - Feature English Name Function
# ============================================================

def feature_name_to_english(name):

    text = str(name)

    text = text.replace(
        "Package_Size_",
        "Package Size: "
    )

    text = text.replace(
        "Temperature_低溫",
        "Temperature: Chilled"
    )

    text = text.replace(
        "Temperature_常溫",
        "Temperature: Ambient"
    )

    for zh, en in REGION_EN.items():

        text = text.replace(
            f"Destination_Region_{zh}",
            f"Region: {en}"
        )

    for zh, en in INDUSTRY_EN.items():

        text = text.replace(
            f"Industry_{zh}",
            f"Industry: {en}"
        )

    text = text.replace(
        "Customer_Type_",
        "Customer Type: "
    )

    replacements = {

        "Monthly_Target_Packages":
            "Monthly Package Volume",

        "Weight_kg":
            "Package Weight",

        "Distance_km":
            "Delivery Distance",

        "Remote_Flag":
            "Remote Delivery",

        "COD_Flag":
            "COD Service",

        "COD_Amount":
            "COD Amount",

        "Insurance_Flag":
            "Insurance Service",

        "Declared_Value":
            "Declared Value",

        "Redelivery_Flag":
            "Redelivery",

        "SameDay_Flag":
            "Same-Day Delivery",

        "StorePickup_Flag":
            "Store Pickup",

        "Pickup_Density":
            "Pickup Density",

        "Customer_Cold_Ratio":
            "Customer Chilled Ratio",

        "Customer_Remote_Ratio":
            "Customer Remote Ratio",

        "Customer_Redelivery_Rate":
            "Customer Redelivery Rate"
    }

    if text in replacements:
        text = replacements[text]

    return text


# ============================================================
# STEP 4 - Main Dashboard
# ============================================================

dashboard_title(

    "物流業 WMS｜XGBoost 成本預測",

    """
    本實作使用與 Multiple Regression、Random Forest
    完全相同的 WMS 資料與測試條件，
    建立 XGBoost Cost-to-Serve 預測模型。
    XGBoost 透過 Boosting 機制，
    讓後續 Decision Trees 持續修正前面模型的預測誤差。
    """
)


# ============================================================
# STEP 5 - Upload Excel
# ============================================================

section_title(

    1,

    "上傳物流業 WMS 教學資料",

    """
    請上傳與 Multiple Regression、Random Forest
    相同的 WMS Excel 資料。
    """
)


uploaded = files.upload()

filename = list(uploaded.keys())[0]


info_box(
    f"已成功載入物流業 WMS 資料：<b>{filename}</b>"
)


# ============================================================
# STEP 6 - Read Data
# ============================================================

section_title(

    2,

    "讀取 WMS Package Data",

    """
    每一列代表一件物流包裹，
    使用相同資料才能公平比較三種演算法。
    """
)


df = pd.read_excel(
    filename,
    sheet_name="WMS_Package_Data"
)


df = df.loc[
    :,
    ~df.columns.astype(str).str.startswith("Unnamed")
]


if "index" in df.columns:

    df = df.drop(
        columns=["index"]
    )


metric_cards([

    (
        "WMS Packages",
        f"{len(df):,}",
        "每列 = 1 件包裹"
    ),

    (
        "Customers",
        f"{df['Customer_ID'].nunique():,}",
        "多客戶物流資料"
    ),

    (
        "Columns",
        f"{len(df.columns)}",
        "WMS Features"
    ),

    (
        "Target",
        "Cost",
        "Actual Operational Cost"
    )
])


# ============================================================
# STEP 7 - Package Size Distribution
# ============================================================

section_title(

    3,

    "觀察 Package Size 分布",

    """
    分析 S60、S90、S120、S150 的資料結構，
    確認不同材積具有足夠的學習樣本。
    """
)


size_counts = (

    df["Package_Size"]

    .value_counts()

    .reindex(
        ["S60", "S90", "S120", "S150"]
    )

    .fillna(0)

    .astype(int)
)


size_percent = (
    size_counts
    / size_counts.sum()
    * 100
)


size_table = pd.DataFrame({

    "Package Size":
        size_counts.index,

    "Package Count":
        size_counts.values,

    "Percentage (%)":
        size_percent.values
})


display(

    warm_style(

        size_table,

        formats={
            "Package Count": "{:,.0f}",
            "Percentage (%)": "{:.2f}%"
        }
    )
)


# ------------------------------------------------------------
# Chart 1
# ------------------------------------------------------------

fig, ax = plt.subplots(
    figsize=(9, 5.5)
)


bars = ax.bar(

    size_counts.index,

    size_counts.values,

    color=[
        SIZE_COLORS[x]
        for x in size_counts.index
    ],

    edgecolor="white"
)


ax.set_title(
    "Package Size Distribution",
    color=COLORS["brown_dark"]
)

ax.set_xlabel("Package Size")
ax.set_ylabel("Number of Packages")

ax.set_facecolor(
    COLORS["cream"]
)

ax.grid(
    axis="y",
    alpha=0.15
)


for bar, value in zip(
    bars,
    size_counts.values
):

    ax.text(

        bar.get_x()
        + bar.get_width()/2,

        value
        + max(size_counts.values)*0.015,

        f"{value:,}",

        ha="center",

        fontweight="bold"
    )


plt.tight_layout()
plt.show()


# ============================================================
# STEP 8 - Define X and y
# ============================================================

section_title(

    4,

    "建立 XGBoost 的 X 與 y",

    """
    X 與 y 必須與前兩個模型完全相同，
    讓最後的差異只來自演算法。
    """
)


numeric_features = [

    "Monthly_Target_Packages",

    "Weight_kg",

    "Distance_km",

    "Remote_Flag",

    "COD_Flag",

    "COD_Amount",

    "Insurance_Flag",

    "Declared_Value",

    "Redelivery_Flag",

    "SameDay_Flag",

    "StorePickup_Flag",

    "Pickup_Density",

    "Customer_Cold_Ratio",

    "Customer_Remote_Ratio",

    "Customer_Redelivery_Rate"
]


categorical_features = [

    "Package_Size",

    "Temperature",

    "Destination_Region",

    "Customer_Type",

    "Industry"
]


required_columns = (

    numeric_features

    + categorical_features

    + [
        "Actual_Operational_Cost",
        "Customer_ID"
    ]
)


missing_columns = [

    col
    for col in required_columns
    if col not in df.columns
]


if len(missing_columns) > 0:

    raise ValueError(
        "Excel 缺少以下欄位：" +
        str(missing_columns)
    )


X = df[
    numeric_features
    + categorical_features
].copy()


y = df[
    "Actual_Operational_Cost"
].copy()


info_box(

    """
    預測目標仍然為 Actual_Operational_Cost。
    Current Quote、企業報價與毛利資料不放入模型，
    避免模型直接學習過去售價。
    """
)


# ============================================================
# STEP 9 - Data Preprocessing
# ============================================================

section_title(

    5,

    "資料前處理",

    """
    數值欄位處理缺值，
    類別欄位透過 One-Hot Encoding 轉換。
    """
)


numeric_transformer = Pipeline(

    steps=[

        (
            "imputer",

            SimpleImputer(
                strategy="median"
            )
        )
    ]
)


categorical_transformer = Pipeline(

    steps=[

        (
            "imputer",

            SimpleImputer(
                strategy="most_frequent"
            )
        ),

        (
            "onehot",

            OneHotEncoder(
                handle_unknown="ignore"
            )
        )
    ]
)


preprocessor = ColumnTransformer(

    transformers=[

        (
            "num",

            numeric_transformer,

            numeric_features
        ),

        (
            "cat",

            categorical_transformer,

            categorical_features
        )
    ]
)


# ============================================================
# STEP 10 - Train / Test Split
# ============================================================

section_title(

    6,

    "依 Customer ID 切分 Training / Testing",

    """
    使用與 Multiple Regression、Random Forest
    相同的 test_size=0.25 與 random_state=42，
    讓三個模型面對相同的新客戶測試資料。
    """
)


groups = df[
    "Customer_ID"
]


splitter = GroupShuffleSplit(

    n_splits=1,

    test_size=0.25,

    random_state=42
)


train_index, test_index = next(

    splitter.split(
        X,
        y,
        groups
    )
)


X_train = X.iloc[
    train_index
].copy()


X_test = X.iloc[
    test_index
].copy()


y_train = y.iloc[
    train_index
].copy()


y_test = y.iloc[
    test_index
].copy()


train_customer_list = sorted(

    df.iloc[
        train_index
    ]["Customer_ID"].unique()
)


test_customer_list = sorted(

    df.iloc[
        test_index
    ]["Customer_ID"].unique()
)


metric_cards([

    (
        "Training Data",

        f"{len(X_train):,}",

        "模型學習資料"
    ),

    (
        "Testing Data",

        f"{len(X_test):,}",

        "模型測試資料"
    ),

    (
        "Training Customers",

        f"{len(train_customer_list)}",

        "訓練客戶"
    ),

    (
        "Testing Customers",

        f"{len(test_customer_list)}",

        "未見新客戶"
    )
])


# ============================================================
# STEP 11 - Build XGBoost
# ============================================================

section_title(

    7,

    "建立 XGBoost 模型",

    """
    XGBoost 使用 Boosting：
    第一棵樹先進行預測，
    後續的樹持續學習前面模型尚未預測好的誤差。
    """
)


model_xgb = Pipeline(

    steps=[

        (
            "preprocessor",

            preprocessor
        ),

        (
            "xgboost",

            XGBRegressor(

                # 決策樹數量
                n_estimators=500,

                # 每棵樹的學習速度
                learning_rate=0.05,

                # 每棵樹最大深度
                max_depth=5,

                # 每棵樹使用 85% Training Data
                subsample=0.85,

                # 每棵樹使用 85% Features
                colsample_bytree=0.85,

                # L2 Regularization
                reg_lambda=1.0,

                # L1 Regularization
                reg_alpha=0.0,

                # Regression Objective
                objective="reg:squarederror",

                # 評估指標
                eval_metric="rmse",

                random_state=42,

                n_jobs=-1
            )
        )
    ]
)


# ============================================================
# STEP 12 - Train XGBoost
# ============================================================

model_xgb.fit(
    X_train,
    y_train
)


info_box(

    """
    XGBoost 模型訓練完成。
    本模型使用 500 棵 Boosting Trees，
    learning_rate = 0.05，
    每棵新樹持續修正前面模型未預測好的部分。
    """
)


# ============================================================
# STEP 13 - Prediction
# ============================================================

pred_xgb = model_xgb.predict(
    X_test
)


# ============================================================
# STEP 14 - Evaluation
# ============================================================

r2 = r2_score(
    y_test,
    pred_xgb
)


mae = mean_absolute_error(
    y_test,
    pred_xgb
)


rmse = np.sqrt(

    mean_squared_error(
        y_test,
        pred_xgb
    )
)


mape = (

    mean_absolute_percentage_error(
        y_test,
        pred_xgb
    )

    * 100
)


section_title(

    8,

    "XGBoost 模型績效",

    """
    使用與前兩個模型相同的
    R²、MAE、RMSE、MAPE 評估。
    """
)


metric_cards([

    (
        "R²",

        f"{r2:.4f}",

        "越接近 1 越好"
    ),

    (
        "MAE",

        f"${mae:.2f}",

        "平均成本預測誤差"
    ),

    (
        "RMSE",

        f"${rmse:.2f}",

        "大型預測誤差"
    ),

    (
        "MAPE",

        f"{mape:.2f}%",

        "平均百分比誤差"
    )
])


# ============================================================
# STEP 15 - Prediction Result
# ============================================================

result_xgb = df.iloc[
    test_index
][
    [
        "Package_ID",
        "Customer_ID",
        "Industry",
        "Customer_Type",
        "Monthly_Target_Packages",
        "Package_Size",
        "Weight_kg",
        "Temperature",
        "Destination_Region",
        "Remote_Flag",
        "Distance_km",
        "COD_Flag",
        "COD_Amount",
        "Redelivery_Flag",
        "SameDay_Flag",
        "Pickup_Density",
        "Actual_Operational_Cost"
    ]
].copy()


result_xgb[
    "XGB_Predicted_Cost"
] = pred_xgb


result_xgb[
    "Prediction_Error"
] = (

    result_xgb[
        "XGB_Predicted_Cost"
    ]

    -

    result_xgb[
        "Actual_Operational_Cost"
    ]
)


result_xgb[
    "Absolute_Error"
] = abs(

    result_xgb[
        "Prediction_Error"
    ]
)


result_xgb[
    "Percentage_Error"
] = (

    result_xgb[
        "Absolute_Error"
    ]

    /

    result_xgb[
        "Actual_Operational_Cost"
    ]

    * 100
)


section_title(

    9,

    "XGBoost 成本預測結果",

    """
    逐筆比較 Actual Cost
    與 XGBoost Predicted Cost。
    """
)


display(

    warm_style(

        result_xgb.head(20),

        formats={

            "Weight_kg": "{:.2f}",

            "Distance_km": "{:.1f}",

            "Actual_Operational_Cost": "${:.2f}",

            "XGB_Predicted_Cost": "${:.2f}",

            "Prediction_Error": "${:.2f}",

            "Absolute_Error": "${:.2f}",

            "Percentage_Error": "{:.2f}%"
        },

        gradient_cols=[
            "Absolute_Error",
            "Percentage_Error"
        ]
    )
)


# ============================================================
# STEP 16 - Actual vs Predicted
# ============================================================

section_title(

    10,

    "Actual Cost vs Predicted Cost",

    """
    不同 Package Size 使用不同顏色，
    越靠近對角線表示預測越準確。
    """
)


fig, ax = plt.subplots(
    figsize=(9, 7)
)


for size in [

    "S60",
    "S90",
    "S120",
    "S150"
]:

    temp = result_xgb[

        result_xgb[
            "Package_Size"
        ] == size
    ]


    ax.scatter(

        temp[
            "Actual_Operational_Cost"
        ],

        temp[
            "XGB_Predicted_Cost"
        ],

        label=size,

        alpha=0.65,

        s=45,

        color=SIZE_COLORS[size],

        edgecolor="white",

        linewidth=0.5
    )


min_value = min(

    result_xgb[
        "Actual_Operational_Cost"
    ].min(),

    result_xgb[
        "XGB_Predicted_Cost"
    ].min()
)


max_value = max(

    result_xgb[
        "Actual_Operational_Cost"
    ].max(),

    result_xgb[
        "XGB_Predicted_Cost"
    ].max()
)


ax.plot(

    [min_value, max_value],

    [min_value, max_value],

    color=COLORS["brown"],

    linestyle="--",

    linewidth=2,

    label="Perfect Prediction"
)


ax.set_title(

    "XGBoost\nActual Cost vs Predicted Cost",

    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Actual Operational Cost"
)


ax.set_ylabel(
    "XGBoost Predicted Cost"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    alpha=0.15
)


ax.legend(
    frameon=False,
    title="Package Size"
)


plt.tight_layout()
plt.show()


# ============================================================
# STEP 17 - Package Size Error
# ============================================================

size_error = (

    result_xgb

    .groupby(
        "Package_Size"
    )

    .agg(

        Package_Count=(
            "Package_ID",
            "count"
        ),

        Actual_Avg_Cost=(
            "Actual_Operational_Cost",
            "mean"
        ),

        Predicted_Avg_Cost=(
            "XGB_Predicted_Cost",
            "mean"
        ),

        MAE=(
            "Absolute_Error",
            "mean"
        ),

        MAPE=(
            "Percentage_Error",
            "mean"
        )
    )

    .reindex(
        ["S60", "S90", "S120", "S150"]
    )

    .reset_index()

    .round(2)
)


section_title(

    11,

    "不同 Package Size 的預測誤差",

    """
    比較 XGBoost 在 S60、S90、
    S120、S150 上的預測表現。
    """
)


display(

    warm_style(

        size_error,

        formats={

            "Package_Count": "{:,.0f}",

            "Actual_Avg_Cost": "${:.2f}",

            "Predicted_Avg_Cost": "${:.2f}",

            "MAE": "${:.2f}",

            "MAPE": "{:.2f}%"
        },

        gradient_cols=[
            "MAE",
            "MAPE"
        ]
    )
)


# ------------------------------------------------------------
# Chart
# ------------------------------------------------------------

fig, ax = plt.subplots(
    figsize=(9, 5.5)
)


bars = ax.bar(

    size_error[
        "Package_Size"
    ],

    size_error[
        "MAE"
    ],

    color=[

        SIZE_COLORS[x]

        for x in

        size_error[
            "Package_Size"
        ]
    ],

    edgecolor="white"
)


ax.set_title(
    "XGBoost Prediction Error by Package Size",
    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Package Size"
)


ax.set_ylabel(
    "Mean Absolute Error"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


for bar, value in zip(
    bars,
    size_error["MAE"]
):

    ax.text(

        bar.get_x()
        + bar.get_width()/2,

        value + 0.1,

        f"${value:.2f}",

        ha="center",

        fontweight="bold"
    )


plt.tight_layout()
plt.show()


# ============================================================
# STEP 18 - Temperature Analysis
# ============================================================

temperature_error = (

    result_xgb

    .groupby(
        "Temperature"
    )

    .agg(

        Package_Count=(
            "Package_ID",
            "count"
        ),

        Actual_Avg_Cost=(
            "Actual_Operational_Cost",
            "mean"
        ),

        Predicted_Avg_Cost=(
            "XGB_Predicted_Cost",
            "mean"
        ),

        MAE=(
            "Absolute_Error",
            "mean"
        )
    )

    .reset_index()

    .round(2)
)


section_title(

    12,

    "常溫與低溫成本預測比較",

    """
    比較 XGBoost 在不同 Temperature
    情境下的 Cost-to-Serve 預測表現。
    """
)


display(

    warm_style(

        temperature_error,

        formats={

            "Package_Count": "{:,.0f}",

            "Actual_Avg_Cost": "${:.2f}",

            "Predicted_Avg_Cost": "${:.2f}",

            "MAE": "${:.2f}"
        },

        gradient_cols=[
            "MAE"
        ]
    )
)


temperature_plot = (
    temperature_error.copy()
)


temperature_plot[
    "Temperature_EN"
] = (

    temperature_plot[
        "Temperature"
    ]

    .map(
        TEMPERATURE_EN
    )

    .fillna(
        temperature_plot[
            "Temperature"
        ]
    )
)


fig, ax = plt.subplots(
    figsize=(7.5, 5)
)


bars = ax.bar(

    temperature_plot[
        "Temperature_EN"
    ],

    temperature_plot[
        "MAE"
    ],

    color=[

        TEMP_COLORS.get(
            x,
            COLORS["peach"]
        )

        for x in

        temperature_plot[
            "Temperature_EN"
        ]
    ],

    edgecolor="white"
)


ax.set_title(
    "XGBoost Prediction Error by Temperature",
    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Temperature Type"
)


ax.set_ylabel(
    "Mean Absolute Error"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


plt.tight_layout()
plt.show()


# ============================================================
# STEP 19 - Region Analysis
# ============================================================

region_error = (

    result_xgb

    .groupby(
        "Destination_Region"
    )

    .agg(

        Package_Count=(
            "Package_ID",
            "count"
        ),

        Actual_Avg_Cost=(
            "Actual_Operational_Cost",
            "mean"
        ),

        Predicted_Avg_Cost=(
            "XGB_Predicted_Cost",
            "mean"
        ),

        MAE=(
            "Absolute_Error",
            "mean"
        )
    )

    .reset_index()

    .round(2)
)


section_title(

    13,

    "不同配送區域成本預測比較",

    """
    比較不同 Destination Region
    的 Cost-to-Serve 預測誤差。
    """
)


display(

    warm_style(

        region_error,

        formats={

            "Package_Count": "{:,.0f}",

            "Actual_Avg_Cost": "${:.2f}",

            "Predicted_Avg_Cost": "${:.2f}",

            "MAE": "${:.2f}"
        },

        gradient_cols=[
            "MAE"
        ]
    )
)


region_plot = (
    region_error.copy()
)


region_plot[
    "Region_EN"
] = (

    region_plot[
        "Destination_Region"
    ]

    .map(
        REGION_EN
    )

    .fillna(
        region_plot[
            "Destination_Region"
        ]
    )
)


region_order = [

    "North",
    "Central",
    "South",
    "East",
    "Offshore"
]


region_plot[
    "Region_EN"
] = pd.Categorical(

    region_plot[
        "Region_EN"
    ],

    categories=region_order,

    ordered=True
)


region_plot = (

    region_plot

    .sort_values(
        "Region_EN"
    )
)


fig, ax = plt.subplots(
    figsize=(9, 5.5)
)


bars = ax.bar(

    region_plot[
        "Region_EN"
    ].astype(str),

    region_plot[
        "MAE"
    ],

    color=[

        REGION_COLORS.get(
            str(x),
            COLORS["peach"]
        )

        for x in

        region_plot[
            "Region_EN"
        ]
    ],

    edgecolor="white"
)


ax.set_title(
    "XGBoost Prediction Error by Destination Region",
    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Destination Region"
)


ax.set_ylabel(
    "Mean Absolute Error"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


plt.tight_layout()
plt.show()


# ============================================================
# STEP 20 - Worst Cases
# ============================================================

worst_cases = (

    result_xgb

    .sort_values(

        "Absolute_Error",

        ascending=False
    )

    .head(20)
)


section_title(

    14,

    "XGBoost 預測誤差最大的 20 件包裹",

    """
    找出 XGBoost 最難預測的物流情境，
    觀察模型剩餘的主要誤差來源。
    """
)


display(

    warm_style(

        worst_cases,

        formats={

            "Weight_kg": "{:.2f}",

            "Distance_km": "{:.1f}",

            "Actual_Operational_Cost": "${:.2f}",

            "XGB_Predicted_Cost": "${:.2f}",

            "Prediction_Error": "${:.2f}",

            "Absolute_Error": "${:.2f}",

            "Percentage_Error": "{:.2f}%"
        },

        gradient_cols=[
            "Absolute_Error",
            "Percentage_Error"
        ]
    )
)


# ============================================================
# STEP 21 - Feature Importance
# ============================================================

section_title(

    15,

    "XGBoost Feature Importance",

    """
    分析 XGBoost 在成本預測時，
    哪些物流特徵具有較高的重要性。
    """
)


fitted_preprocessor = (

    model_xgb

    .named_steps[
        "preprocessor"
    ]
)


feature_names = (

    fitted_preprocessor

    .get_feature_names_out()
)


feature_names = [

    str(x)
    .replace("num__", "")
    .replace("cat__", "")

    for x in feature_names
]


importance_values = (

    model_xgb

    .named_steps[
        "xgboost"
    ]

    .feature_importances_
)


importance_df = pd.DataFrame({

    "Feature":
        feature_names,

    "Importance":
        importance_values
})


importance_df[
    "Feature_EN"
] = (

    importance_df[
        "Feature"
    ]

    .apply(
        feature_name_to_english
    )
)


importance_df = (

    importance_df

    .sort_values(

        "Importance",

        ascending=False
    )
)


top_importance = (

    importance_df

    .head(20)

    .copy()
)


display(

    warm_style(

        top_importance[
            [
                "Feature",
                "Importance"
            ]
        ],

        formats={
            "Importance": "{:.4f}"
        },

        gradient_cols=[
            "Importance"
        ]
    )
)


# ============================================================
# Chart - Feature Importance
# ============================================================

importance_chart = (

    top_importance

    .sort_values(
        "Importance"
    )
)


fig, ax = plt.subplots(
    figsize=(11, 8)
)


ax.barh(

    importance_chart[
        "Feature_EN"
    ],

    importance_chart[
        "Importance"
    ],

    color=COLORS["coral"]
)


ax.set_title(

    "Top Cost Drivers\nXGBoost Feature Importance",

    color=COLORS["brown_dark"],

    pad=15
)


ax.set_xlabel(
    "Feature Importance"
)


ax.set_ylabel(
    "Cost Driver"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="x",
    alpha=0.15
)


plt.tight_layout()
plt.show()


# ============================================================
# STEP 22 - Model Summary
# ============================================================

model_summary = pd.DataFrame({

    "Model": [
        "XGBoost"
    ],

    "R2": [
        r2
    ],

    "MAE_Cost": [
        mae
    ],

    "RMSE_Cost": [
        rmse
    ],

    "MAPE_Cost_%": [
        mape
    ],

    "Trees": [
        500
    ],

    "Learning_Rate": [
        0.05
    ],

    "Max_Depth": [
        5
    ],

    "Training_Records": [
        len(X_train)
    ],

    "Testing_Records": [
        len(X_test)
    ],

    "Training_Customers": [
        len(train_customer_list)
    ],

    "Testing_Customers": [
        len(test_customer_list)
    ]
})


section_title(

    16,

    "XGBoost 模型總結",

    """
    完成第三個 Cost-to-Serve 預測模型。
    下一階段即可正式進行
    Multiple Regression、Random Forest、
    XGBoost 三模型比較。
    """
)


display(

    warm_style(

        model_summary,

        formats={

            "R2": "{:.4f}",

            "MAE_Cost": "${:.2f}",

            "RMSE_Cost": "${:.2f}",

            "MAPE_Cost_%": "{:.2f}%",

            "Learning_Rate": "{:.2f}"
        },

        gradient_cols=[
            "R2"
        ]
    )
)


# ============================================================
# STEP 23 - Export Excel
# ============================================================

output_filename = (
    "Logistics_WMS_XGBoost_Cost_Prediction.xlsx"
)


with pd.ExcelWriter(

    output_filename,

    engine="openpyxl"

) as writer:


    model_summary.to_excel(
        writer,
        sheet_name="Model_Summary",
        index=False
    )


    result_xgb.to_excel(
        writer,
        sheet_name="Package_Predictions",
        index=False
    )


    worst_cases.to_excel(
        writer,
        sheet_name="Worst_Cases",
        index=False
    )


    importance_df.to_excel(
        writer,
        sheet_name="Feature_Importance",
        index=False
    )


    size_error.to_excel(
        writer,
        sheet_name="Size_Analysis",
        index=False
    )


    temperature_error.to_excel(
        writer,
        sheet_name="Temperature_Analysis",
        index=False
    )


    region_error.to_excel(
        writer,
        sheet_name="Region_Analysis",
        index=False
    )


# ============================================================
# STEP 24 - Style Excel
# ============================================================

from openpyxl import load_workbook

from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)


wb = load_workbook(
    output_filename
)


header_fill = PatternFill(
    "solid",
    fgColor="C96F5B"
)


alternate_fill = PatternFill(
    "solid",
    fgColor="FFF8F0"
)


header_font = Font(
    color="FFFFFF",
    bold=True
)


body_font = Font(
    color="4F3A30"
)


thin_border = Border(

    bottom=Side(
        style="thin",
        color="EBDDD3"
    )
)


for ws in wb.worksheets:


    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions


    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    for row in range(
        2,
        ws.max_row + 1
    ):

        if row % 2 == 0:

            for cell in ws[row]:

                cell.fill = alternate_fill


        for cell in ws[row]:

            cell.font = body_font

            cell.border = thin_border

            cell.alignment = Alignment(
                vertical="center"
            )


    for column_cells in ws.columns:


        length = 0

        column_letter = (
            column_cells[0]
            .column_letter
        )


        for cell in column_cells:

            try:

                value = str(
                    cell.value
                    if cell.value is not None
                    else ""
                )

                length = max(
                    length,
                    len(value)
                )

            except:
                pass


        ws.column_dimensions[
            column_letter
        ].width = min(
            max(length + 3, 12),
            32
        )


    ws.row_dimensions[1].height = 24


wb.save(
    output_filename
)


# ============================================================
# STEP 25 - Final Dashboard
# ============================================================

dashboard_title(

    "物流業 WMS｜XGBoost 分析完成",

    """
    已完成 XGBoost Cost-to-Serve 成本預測，
    包含模型 KPI、Package Size、
    Temperature、Destination Region、
    Worst Cases 與 Feature Importance 分析。
    """
)


metric_cards([

    (
        "R²",

        f"{r2:.4f}",

        "模型解釋能力"
    ),

    (
        "MAE",

        f"${mae:.2f}",

        "平均成本誤差"
    ),

    (
        "RMSE",

        f"${rmse:.2f}",

        "大型誤差風險"
    ),

    (
        "MAPE",

        f"{mape:.2f}%",

        "平均百分比誤差"
    )
])


info_box(

    """
    至此已完成 Multiple Regression、
    Random Forest 與 XGBoost 三個模型。
    下一步應將三種模型的
    R²、MAE、RMSE、MAPE 放在同一張表與圖中比較，
    再由實際測試結果選出最佳 Cost-to-Serve 模型。
    """
)


print(
    "📁 Analysis file:",
    output_filename
)


files.download(
    output_filename
)
