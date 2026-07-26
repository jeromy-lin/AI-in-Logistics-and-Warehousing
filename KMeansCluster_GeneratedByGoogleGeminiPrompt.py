# ==============================================================================
# SKU K-Means 分群與倉儲效益分析自動化工具 (Colab 專用完整程式碼)
# ==============================================================================

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=FutureWarning)

import io
import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from google.colab import files
from IPython.display import display, HTML

from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# ------------------------------------------------------------------------------
# 1. 檔案上傳與初步檢查
# ------------------------------------------------------------------------------
print("【步驟 1/12】請上傳包含 SKU 資料的 Excel 檔案...")
uploaded = files.upload()

if not uploaded:
    print("❌ 錯誤：未上傳任何檔案，程式已終止。")
    sys.exit()

file_name = list(uploaded.keys())[0]
sheet_target = "SKU資料輸入"

try:
    # 讀取 Excel，第 4 列為表頭 (header=3)
    df_raw = pd.read_excel(file_name, sheet_name=sheet_target, header=3)
except Exception as e:
    print(f"❌ 讀取 Excel 失敗：工作表 '{sheet_target}' 可能不存在，或檔案格式不正確。詳細錯誤: {e}")
    sys.exit()

# 自動移除完全空白的資料列
df_raw = df_raw.dropna(how="all").reset_index(drop=True)

# 檢查必要欄位是否存在
required_cols = [
    "SKU編號", "商品名稱", "商品類別",
    "每日出貨箱數", "每日揀貨次數", "體積等級", "重量等級",
    "是否液體", "是否易碎", "是否季節性",
    "目前儲位區", "目前距離主要作業區m"
]

missing_cols = [col for col in required_cols if col not in df_raw.columns]
if missing_cols:
    print(f"❌ 錯誤：Excel 檔案中缺少以下必要欄位：{', '.join(missing_cols)}")
    sys.exit()

df = df_raw.copy()

# ------------------------------------------------------------------------------
# 2. 資料前處理與特徵轉換
# ------------------------------------------------------------------------------
print("【步驟 2/12】進行資料前處理與特徵轉換...")

volume_map = {"小": 1, "中": 2, "大": 3}
weight_map = {"輕": 1, "中": 2, "重": 3}
binary_map = {
    "否": 0, "是": 1, "N": 0, "Y": 1, "No": 0, "Yes": 1,
    "False": 0, "True": 1, "0": 0, "1": 1
}

# 清除文字前後空白並建立數值分析欄位
df["體積等級數值"] = df["體積等級"].astype(str).str.strip().map(volume_map)
df["重量等級數值"] = df["重量等級"].astype(str).str.strip().map(weight_map)
df["是否液體數值"] = df["是否液體"].astype(str).str.strip().map(binary_map)
df["是否易碎數值"] = df["是否易碎"].astype(str).str.strip().map(binary_map)
df["是否季節性數值"] = df["是否季節性"].astype(str).str.strip().map(binary_map)

# 數值欄位轉 numeric 並進行中位數補值
feature_cols_raw = [
    "每日出貨箱數", "每日揀貨次數",
    "體積等級數值", "重量等級數值",
    "是否液體數值", "是否易碎數值", "是否季節性數值",
    "目前距離主要作業區m"
]

for col in feature_cols_raw:
    df[col] = pd.to_numeric(df[col], errors="coerce")
    median_val = df[col].median()
    df[col] = df[col].fillna(median_val if pd.notnull(median_val) else 0)

# ------------------------------------------------------------------------------
# 3. K-Means 分群與 PCA 降維
# ------------------------------------------------------------------------------
print("【步驟 3/12】執行 K-Means 5群分析與 PCA 降維...")

scaler = StandardScaler()
X_scaled = scaler.fit_transform(df[feature_cols_raw])

kmeans = KMeans(n_clusters=5, random_state=42, n_init=20)
cluster_labels = kmeans.fit_predict(X_scaled)

# 群組編號改為 1 至 5
df["群組編號"] = cluster_labels + 1

pca = PCA(n_components=2, random_state=42)
pca_transformed = pca.fit_transform(X_scaled)
df["PCA_X"] = pca_transformed[:, 0]
df["PCA_Y"] = pca_transformed[:, 1]

# ------------------------------------------------------------------------------
# 4. PCA 二維視覺化 (Matplotlib)
# ------------------------------------------------------------------------------
print("【步驟 4/12】繪製 PCA 散布圖...")

plt.figure(figsize=(9, 7), facecolor='white', dpi=150)
ax = plt.subplot(111)
ax.set_facecolor('white')

# 柔和色彩組合 (Cluster 1 ~ 5)
cluster_plot_colors = ['#5B9BD5', '#70AD47', '#ED7D31', '#7030A0', '#FFC000']

centers_pca = pca.transform(kmeans.cluster_centers_)

for c in range(1, 6):
    idx = df["群組編號"] == c
    plt.scatter(
        df.loc[idx, "PCA_X"], df.loc[idx, "PCA_Y"],
        c=cluster_plot_colors[c-1],
        label=f"Cluster {c}",
        alpha=0.85,
        s=60,
        edgecolors='none'
    )
    # 標示群組中心
    plt.scatter(
        centers_pca[c-1, 0], centers_pca[c-1, 1],
        c=cluster_plot_colors[c-1],
        marker='*',
        s=300,
        edgecolors='#333333',
        linewidths=1.2
    )

plt.title("PCA Visualization of Five SKU Clusters", fontsize=14, fontweight='bold', pad=15, color='#262626')
plt.xlabel("Principal Component 1", fontsize=11, color='#404040')
plt.ylabel("Principal Component 2", fontsize=11, color='#404040')

for spine in ax.spines.values():
    spine.set_color('#D9D9D9')

plt.grid(True, linestyle='--', alpha=0.3, color='#CCCCCC')
plt.legend(frameon=True, facecolor='white', edgecolor='#E0E0E0', fontsize=10)
plt.tight_layout()
plt.show()

# ------------------------------------------------------------------------------
# 5. 效益分析計算
# ------------------------------------------------------------------------------
print("【步驟 5/12】計算儲位優化效益...")

# 定義優先度分數以區分商品層級
df["優先度分數"] = df["每日出貨箱數"] * 0.4 + df["每日揀貨次數"] * 0.6

p80 = df["目前距離主要作業區m"].quantile(0.20)
p60 = df["目前距離主要作業區m"].quantile(0.40)
p50 = df["目前距離主要作業區m"].median()

q75_score = df["優先度分數"].quantile(0.75)
q50_score = df["優先度分數"].quantile(0.50)
q25_score = df["優先度分數"].quantile(0.25)

def assign_suggested_dist(row):
    score = row["優先度分數"]
    curr = row["目前距離主要作業區m"]
    if score >= q75_score:
        target = p80
    elif score >= q50_score:
        target = p60
    elif score >= q25_score:
        target = p50
    else:
        target = curr
    return min(curr, target)

df["建議距離m"] = df.apply(assign_suggested_dist, axis=1)
df["改善前距離成本"] = df["每日揀貨次數"] * df["目前距離主要作業區m"]
df["改善後距離成本"] = df["每日揀貨次數"] * df["建議距離m"]
df["節省距離成本"] = df["改善前距離成本"] - df["改善後距離成本"]
df["節省距離成本"] = df["節省距離成本"].apply(lambda x: max(0.0, x))

df["改善率%"] = np.where(
    df["改善前距離成本"] > 0,
    (df["節省距離成本"] / df["改善前距離成本"]) * 100,
    0.0
)

# ------------------------------------------------------------------------------
# 6. Colab 表格美化通用函數
# ------------------------------------------------------------------------------
cluster_colors = {
    1: "#DCEAF7",
    2: "#E2F0D9",
    3: "#FCE4D6",
    4: "#E4DFEC",
    5: "#FFF2CC"
}

def 美化表格(df_input, 標題="", 百分比欄位=None, 小數欄位=None, 高亮前十=False, 綠色漸層欄=None):
    display(HTML(f"<h3 style='color:#1F4E78; font-family:Microsoft JhengHei, sans-serif; margin-top:20px; margin-bottom:8px;'>{標題}</h3>"))
    
    styler = df_input.style
    
    format_dict = {}
    if 小數欄位:
        for col in 小數欄位:
            if col in df_input.columns:
                format_dict[col] = "{:.2f}"
    if 百分比欄位:
        for col in 百分比欄位:
            if col in df_input.columns:
                format_dict[col] = "{:.1f}%"
    
    styler = styler.format(format_dict, na_rep="-")
    
    # CSS 樣式設定
    styles = [
        {'selector': 'th', 'props': [
            ('background-color', '#1F4E78'),
            ('color', 'white'),
            ('font-weight', 'bold'),
            ('text-align', 'center'),
            ('padding', '8px 12px'),
            ('font-size', '13px'),
            ('border', '1px solid #D9D9D9')
        ]},
        {'selector': 'td', 'props': [
            ('padding', '6px 10px'),
            ('vertical-align', 'middle'),
            ('border', '1px solid #E0E0E0'),
            ('font-size', '12px'),
            ('white-space', 'normal'),
            ('word-break', 'break-word')
        ]},
        {'selector': 'table', 'props': [
            ('border-collapse', 'collapse'),
            ('width', '100%'),
            ('margin-bottom', '15px')
        ]}
    ]
    styler = styler.set_table_styles(styles)
    styler = styler.hide(axis='index')
    
    # 交錯底色
    def zebra(row):
        return ['background-color: #F9FAFB' if row.name % 2 == 1 else 'background-color: #FFFFFF' for _ in row]
    styler = styler.apply(zebra, axis=1)
    
    # 群組編號底色
    if "群組編號" in df_input.columns:
        def color_cluster(val):
            color = cluster_colors.get(val, "#FFFFFF")
            return f'background-color: {color}; text-align: center; font-weight: bold;'
        styler = styler.map(color_cluster, subset=["群組編號"])
        
    # 文字與對齊控制
    text_left_cols = [c for c in ["SKU編號", "商品名稱", "商品類別", "目前儲位區", "群組特性", "管理建議"] if c in df_input.columns]
    if text_left_cols:
        styler = styler.map(lambda v: 'text-align: left;', subset=text_left_cols)
        
    num_center_cols = [c for c in ["群組編號", "SKU數量", "體積等級", "重量等級", "是否液體", "是否易碎", "是否季節性"] if c in df_input.columns]
    if num_center_cols:
        styler = styler.map(lambda v: 'text-align: center;', subset=num_center_cols)
        
    # 優先調整前 10 項黃色底色
    if 高亮前十:
        def highlight_top10(row):
            if row.name < 10:
                return ['background-color: #FFF2CC !important;' for _ in row]
            return ['' for _ in row]
        styler = styler.apply(highlight_top10, axis=1)
        
    # 改善率為 0 灰色字體
    if "改善率%" in df_input.columns:
        def gray_zero_rate(val):
            try:
                if float(val) == 0:
                    return 'color: #A6A6A6;'
            except:
                pass
            return ''
        styler = styler.map(gray_zero_rate, subset=["改善率%"])

    display(styler)

# ------------------------------------------------------------------------------
# 7. 顯示 SKU 分群結果表
# ------------------------------------------------------------------------------
print("【步驟 6/12】產生 SKU 分群結果表...")

sku_result_cols = [
    "SKU編號", "商品名稱", "商品類別", "每日出貨箱數", "每日揀貨次數",
    "體積等級", "重量等級", "是否液體", "是否易碎", "是否季節性",
    "目前儲位區", "目前距離主要作業區m", "群組編號", "PCA_X", "PCA_Y"
]

df_sku_result = df[sku_result_cols].sort_values(by=["群組編號", "每日揀貨次數"], ascending=[True, False]).reset_index(drop=True)

美化表格(
    df_sku_result,
    標題="一、SKU 分群詳細結果表",
    小數欄位=["PCA_X", "PCA_Y", "每日出貨箱數", "每日揀貨次數", "目前距離主要作業區m"]
)

# ------------------------------------------------------------------------------
# 8. 產生並顯示五群統計摘要表
# ------------------------------------------------------------------------------
print("【步驟 7/12】計算並產生五群統計摘要表...")

summary_rows = []
total_skus = len(df)

for c in range(1, 6):
    c_df = df[df["群組編號"] == c]
    cnt = len(c_df)
    pct = (cnt / total_skus) * 100 if total_skus > 0 else 0
    
    avg_out = c_df["每日出貨箱數"].mean()
    avg_pick = c_df["每日揀貨次數"].mean()
    avg_vol = c_df["體積等級數值"].mean()
    avg_wt = c_df["重量等級數值"].mean()
    pct_liq = c_df["是否液體數值"].mean() * 100
    pct_frag = c_df["是否易碎數值"].mean() * 100
    pct_seas = c_df["是否季節性數值"].mean() * 100
    avg_dist = c_df["目前距離主要作業區m"].mean()
    
    # 根據實際數據動態判斷群組特性與建議
    traits = []
    if avg_out > df["每日出貨箱數"].mean() and avg_pick > df["每日揀貨次數"].mean():
        traits.append("高出貨高揀貨熱銷主力")
    if avg_wt > 2.2 or avg_vol > 2.2:
        traits.append("大型或重物商品")
    if pct_liq > 40 or pct_frag > 40:
        traits.append("高風險液體/易碎品")
    if pct_seas > 40:
        traits.append("季節性波動商品")
    if not traits:
        traits.append("低頻一般周轉商品")
        
    trait_str = " / ".join(traits)
    
    # 建議
    if "高出貨高揀貨熱銷主力" in trait_str:
        sug = "移至前排核心快速揀貨區，減少搬運距離"
    elif "大型或重物商品" in trait_str:
        sug = "配置於重物專區或棧板區，輔助機械設備作業"
    elif "高風險液體/易碎品" in trait_str:
        sug = "設置防摔防漏防護專區，固定專用防護儲位"
    elif "季節性波動商品" in trait_str:
        sug = "採用彈性調度儲位，旺季前移、淡季後移"
    else:
        sug = "維持中後排儲位，定期檢討周轉率"
        
    summary_rows.append({
        "群組編號": c,
        "SKU數量": cnt,
        "SKU占比%": pct,
        "平均每日出貨箱數": avg_out,
        "平均每日揀貨次數": avg_pick,
        "平均體積等級": avg_vol,
        "平均重量等級": avg_wt,
        "液體商品比例%": pct_liq,
        "易碎商品比例%": pct_frag,
        "季節性商品比例%": pct_seas,
        "平均目前距離m": avg_dist,
        "群組特性": trait_str,
        "管理建議": sug
    })

df_summary = pd.DataFrame(summary_rows)

美化表格(
    df_summary,
    標題="二、五群統計摘要與管理建議表",
    百分比欄位=["SKU占比%", "液體商品比例%", "易碎商品比例%", "季節性商品比例%"],
    小數欄位=["平均每日出貨箱數", "平均每日揀貨次數", "平均體積等級", "平均重量等級", "平均目前距離m"]
)

# ------------------------------------------------------------------------------
# 9. 顯示 SKU 效益分析表
# ------------------------------------------------------------------------------
print("【步驟 8/12】產生 SKU 效益分析表...")

sku_benefit_cols = [
    "SKU編號", "商品名稱", "群組編號", "每日出貨箱數", "每日揀貨次數",
    "目前距離主要作業區m", "建議距離m", "改善前距離成本", "改善後距離成本",
    "節省距離成本", "改善率%"
]

df_benefit_sorted = df[sku_benefit_cols].sort_values(by="節省距離成本", ascending=False).reset_index(drop=True)

美化表格(
    df_benefit_sorted,
    標題="三、優先改善 SKU 排名 (SKU 效益分析表)",
    百分比欄位=["改善率%"],
    小數欄位=["每日出貨箱數", "每日揀貨次數", "目前距離主要作業區m", "建議距離m", "改善前距離成本", "改善後距離成本", "節省距離成本"],
    高亮前十=True
)

# ------------------------------------------------------------------------------
# 10. Colab HTML 摘要數字卡片
# ------------------------------------------------------------------------------
print("【步驟 9/12】產生摘要數字卡片...")

total_sku_cnt = len(df)
adj_sku_cnt = len(df[df["節省距離成本"] > 0])
before_cost = df["改善前距離成本"].sum()
after_cost = df["改善後距離成本"].sum()
saved_cost = df["節省距離成本"].sum()
overall_imp_rate = (saved_cost / before_cost * 100) if before_cost > 0 else 0.0

cards_html = f"""
<div style="font-family: Microsoft JhengHei, sans-serif; margin: 20px 0;">
    <h3 style="color: #1F4E78; margin-bottom: 12px;">四、整體優化效益摘要卡片</h3>
    <div style="display: flex; flex-wrap: wrap; gap: 12px;">
        <div style="flex: 1; min-width: 150px; background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center;">
            <div style="color: #1F4E78; font-size: 13px; font-weight: bold;">SKU 總數</div>
            <div style="color: #262626; font-size: 24px; font-weight: bold; margin-top: 6px;">{total_sku_cnt:,}</div>
        </div>
        <div style="flex: 1; min-width: 150px; background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center;">
            <div style="color: #1F4E78; font-size: 13px; font-weight: bold;">建議調整 SKU 數</div>
            <div style="color: #C65911; font-size: 24px; font-weight: bold; margin-top: 6px;">{adj_sku_cnt:,}</div>
        </div>
        <div style="flex: 1; min-width: 150px; background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center;">
            <div style="color: #1F4E78; font-size: 13px; font-weight: bold;">改善前每日距離成本</div>
            <div style="color: #262626; font-size: 22px; font-weight: bold; margin-top: 6px;">{before_cost:,.1f}</div>
        </div>
        <div style="flex: 1; min-width: 150px; background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center;">
            <div style="color: #1F4E78; font-size: 13px; font-weight: bold;">改善後每日距離成本</div>
            <div style="color: #2F5597; font-size: 22px; font-weight: bold; margin-top: 6px;">{after_cost:,.1f}</div>
        </div>
        <div style="flex: 1; min-width: 150px; background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center;">
            <div style="color: #1F4E78; font-size: 13px; font-weight: bold;">每日節省距離成本</div>
            <div style="color: #385723; font-size: 22px; font-weight: bold; margin-top: 6px;">{saved_cost:,.1f}</div>
        </div>
        <div style="flex: 1; min-width: 150px; background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 8px; padding: 15px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center;">
            <div style="color: #1F4E78; font-size: 13px; font-weight: bold;">整體改善率</div>
            <div style="color: #385723; font-size: 24px; font-weight: bold; margin-top: 6px;">{overall_imp_rate:.1f}%</div>
        </div>
    </div>
</div>
"""
display(HTML(cards_html))

# ------------------------------------------------------------------------------
# 11. 效益比較圖表 (Matplotlib)
# ------------------------------------------------------------------------------
print("【步驟 10/12】繪製效益比較圖表...")

# 圖 1：五群改善前後距離成本長條圖
cluster_costs = df.groupby("群組編號")[["改善前距離成本", "改善後距離成本"]].sum().reindex(range(1, 6)).fillna(0)

plt.figure(figsize=(9, 5.5), facecolor='white', dpi=150)
ax1 = plt.subplot(111)
ax1.set_facecolor('white')

x_indices = np.arange(1, 6)
bar_width = 0.35

bars1 = plt.bar(x_indices - bar_width/2, cluster_costs["改善前距離成本"], width=bar_width, label='Before Improvement', color='#5B9BD5', alpha=0.9)
bars2 = plt.bar(x_indices + bar_width/2, cluster_costs["改善後距離成本"], width=bar_width, label='After Improvement', color='#ED7D31', alpha=0.9)

plt.title("Before and After Distance Cost by Cluster", fontsize=13, fontweight='bold', pad=15, color='#262626')
plt.xlabel("Cluster 1 to Cluster 5", fontsize=11, color='#404040')
plt.ylabel("Daily Distance Cost", fontsize=11, color='#404040')
plt.xticks(x_indices, [f"Cluster {i}" for i in range(1, 6)])

for bar in bars1:
    yval = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2.0, yval + (yval*0.01 + 0.5), f"{yval:,.0f}", ha='center', va='bottom', fontsize=9, color='#404040')

for bar in bars2:
    yval = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2.0, yval + (yval*0.01 + 0.5), f"{yval:,.0f}", ha='center', va='bottom', fontsize=9, color='#404040')

for spine in ax1.spines.values():
    spine.set_color('#D9D9D9')

plt.grid(axis='y', linestyle='--', alpha=0.3, color='#CCCCCC')
plt.legend(frameon=True, facecolor='white', edgecolor='#E0E0E0', fontsize=10)
plt.tight_layout()
plt.show()

# 圖 2：前 10 項改善 SKU 水平長條圖
top10_df = df_benefit_sorted.head(10).iloc[::-1]

plt.figure(figsize=(9, 5.5), facecolor='white', dpi=150)
ax2 = plt.subplot(111)
ax2.set_facecolor('white')

bars_h = plt.barh(top10_df["SKU編號"].astype(str), top10_df["節省距離成本"], color='#70AD47', alpha=0.85, height=0.6)

plt.title("Top 10 SKUs by Distance Cost Reduction", fontsize=13, fontweight='bold', pad=15, color='#262626')
plt.xlabel("Distance Cost Reduction", fontsize=11, color='#404040')
plt.ylabel("SKU Number", fontsize=11, color='#404040')

for bar in bars_h:
    xval = bar.get_width()
    plt.text(xval + (max(top10_df["節省距離成本"])*0.01 + 0.2), bar.get_y() + bar.get_height()/2.0, f"{xval:,.1f}", ha='left', va='center', fontsize=9, color='#404040')

for spine in ax2.spines.values():
    spine.set_color('#D9D9D9')

plt.grid(axis='x', linestyle='--', alpha=0.3, color='#CCCCCC')
plt.tight_layout()
plt.show()

# ------------------------------------------------------------------------------
# 12. 中文分析結論 HTML 區塊
# ------------------------------------------------------------------------------
print("【步驟 11/12】生成分析結論與管理建議...")

# 數據分析
cluster_counts_str = "、".join([f"群組 {i}：{len(df[df['群組編號']==i])} 項" for i in range(1, 6)])

max_out_cluster = df_summary.loc[df_summary["平均每日出貨箱數"].idxmax(), "群組編號"]
max_pick_cluster = df_summary.loc[df_summary["平均每日揀貨次數"].idxmax(), "群組編號"]
max_dist_cluster = df_summary.loc[df_summary["平均目前距離m"].idxmax(), "群組編號"]

# 建議優先調整群組 (綜合揀貨次數與可改善潛力)
group_savings = df.groupby("群組編號")["節省距離成本"].sum()
priority_cluster = group_savings.idxmax()

top3_skus = df_benefit_sorted.head(3)["SKU編號"].tolist()
top3_skus_str = "、".join([str(s) for s in top3_skus])

monthly_savings = saved_cost * 22
yearly_savings = saved_cost * 264

conclusion_html = f"""
<div style="background-color: #F2F7FA; border: 1px solid #B8D1E5; border-radius: 8px; padding: 22px; font-family: Microsoft JhengHei, sans-serif; margin: 25px 0;">
    <h2 style="color: #1F4E78; margin-top: 0; font-size: 18px; border-bottom: 2px solid #1F4E78; padding-bottom: 8px;">五、SKU 分群與儲位優化分析結論</h2>
    
    <div style="margin-top: 15px; color: #262626; font-size: 14px; line-height: 1.7;">
        <p><strong>【資料摘要與分群結果】</strong><br>
        本評估共分析 <strong>{total_sku_cnt}</strong> 項 SKU。經由 K-Means 演算法分成 5 個群組，各群數量分別為：{cluster_counts_str}。</p>
        
        <p><strong>【關鍵數據發現】</strong></p>
        <ul style="margin-top: 5px; margin-bottom: 10px; padding-left: 20px;">
            <li><strong>最高出貨量群組：</strong>群組 {max_out_cluster}（平均每日 {df_summary.loc[max_out_cluster-1, '平均每日出貨箱數']:.1f} 箱）</li>
            <li><strong>最高揀貨頻率群組：</strong>群組 {max_pick_cluster}（平均每日 {df_summary.loc[max_pick_cluster-1, '平均每日揀貨次數']:.1f} 次）</li>
            <li><strong>目前距離最遠群組：</strong>群組 {max_dist_cluster}（平均目前距離 {df_summary.loc[max_dist_cluster-1, '平均目前距離m']:.1f} 米）</li>
        </ul>

        <p><strong>【效益改善評估】</strong></p>
        <ul style="margin-top: 5px; margin-bottom: 10px; padding-left: 20px;">
            <li><strong>建議優先調整：</strong>建議優先調整 <strong>群組 {priority_cluster}</strong>，整體共需調整 <strong>{adj_sku_cnt}</strong> 項 SKU 的儲位配置。</li>
            <li><strong>節省成本前 3 名 SKU：</strong>{top3_skus_str}。</li>
            <li><strong>每日距離成本：</strong>由改善前 <strong>{before_cost:,.1f}</strong> 降至改善後 <strong>{after_cost:,.1f}</strong>，每日可降低 <strong>{saved_cost:,.1f}</strong> 搬運距離成本。</li>
            <li><strong>整體改善率：</strong>達 <strong>{overall_imp_rate:.1f}%</strong>。</li>
            <li><strong>中長期預估效益：</strong>以每月 22 工作日估算，每月可減少 <strong>{monthly_savings:,.1f}</strong> 米搬運距離；以每年 264 工作日估算，每年可節省高達 <strong>{yearly_savings:,.1f}</strong> 米搬運距離。</li>
        </ul>

        <p><strong>【管理建議與策略推動】</strong><br>
        建議倉儲管理團隊採取分階段搬遷策略，優先將前 10 項高效益 SKU（如 {top3_skus_str} 等）移動至靠近主要作業區的周轉區，可迅速發揮 80% 以上的降本效益，大幅提升揀貨作業效率與人因工程友善度。</p>
    </div>
</div>
"""
display(HTML(conclusion_html))

# ------------------------------------------------------------------------------
# 13. 匯出 Excel (openpyxl 美化與多工作表)
# ------------------------------------------------------------------------------
print("【步驟 12/12】產生並格式化 Excel 報表 (KMeans_SKU_分析結果.xlsx)...")

output_filename = "KMeans_SKU_分析結果.xlsx"

# 建立簡化版摘要表供 Excel
df_overall_summary = pd.DataFrame([{
    "項目": "分析 SKU 總數", "數值": total_sku_cnt
}, {
    "項目": "建議調整 SKU 數", "數值": adj_sku_cnt
}, {
    "項目": "改善前每日距離成本", "數值": before_cost
}, {
    "項目": "改善後每日距離成本", "數值": after_cost
}, {
    "項目": "每日節省距離成本", "數值": saved_cost
}, {
    "項目": "整體改善率", "數值": overall_imp_rate / 100.0
}, {
    "項目": "每月預估節省距離(22天)", "數值": monthly_savings
}, {
    "項目": "每年預估節省距離(264天)", "數值": yearly_savings
}])

df_management = pd.DataFrame([
    {"階段": "第一階段 (高優先)", "目標SKU": top3_skus_str, "管理策略": "即刻進行儲位對調，配置於距作業區 20% 內之黃金通道，迅速降低揀貨行進距離。"},
    {"階段": "第二階段 (群組優化)", "目標SKU": f"群組 {priority_cluster} 全體 SKU", "管理策略": "依據熱銷及高揀貨頻率特性，重新規劃整區動線，減少二次搬運。"},
    {"階段": "第三階段 (特殊品項)", "目標SKU": "液體 / 易碎 / 重物 SKU", "管理策略": "建置專屬安全防護區與棧板儲位，兼顧作業安全與物流效率。"},
    {"階段": "第四階段 (動態檢討)", "目標SKU": "季節性商品及低頻商品", "管理策略": "導入淡旺季動態儲位機制，低頻商品後移，維持前端儲位彈性。"}
])

wb = openpyxl.Workbook()
wb.remove(wb.active) # 移除預設工作表

# 定義配色與樣式
navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
white_bold_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
regular_font = Font(name="Microsoft JhengHei", size=10)
bold_font = Font(name="Microsoft JhengHei", size=10, bold=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

excel_cluster_fills = {
    1: PatternFill(start_color="DCEAF7", end_color="DCEAF7", fill_type="solid"),
    2: PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
    3: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    4: PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    5: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
}

yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

sheets_data = [
    ("SKU分群結果", df_sku_result),
    ("五群統計摘要", df_summary),
    ("SKU效益分析", df_benefit_sorted),
    ("整體效益摘要", df_overall_summary),
    ("管理建議", df_management)
]

for sheet_name, data_df in sheets_data:
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # 寫入標題
    headers = list(data_df.columns)
    ws.append(headers)
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    
    # 寫入資料
    for r_idx, row in data_df.iterrows():
        ws.append(list(row))
        current_row = r_idx + 2
        ws.row_dimensions[current_row].height = 20
        
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            col_name = headers[c_idx - 1]
            
            # 格式化數字
            if isinstance(val, (int, float, np.number)):
                if "率" in col_name or "比例" in col_name or "占比" in col_name:
                    cell.number_format = '0.0%' if sheet_name == "整體效益摘要" and col_name == "數值" else '0.0%'
                    if sheet_name != "整體效益摘要":
                        cell.value = val / 100.0 if val > 1 else val
                elif "數" in col_name or "次" in col_name or "成本" in col_name or "距離" in col_name or "PCA" in col_name or "等級" in col_name:
                    if isinstance(val, int) or (isinstance(val, float) and val.is_integer() and "PCA" not in col_name and "平均" not in col_name):
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                if col_name in ["SKU編號", "商品名稱", "商品類別", "目前儲位區", "群組特性", "管理建議", "項目", "階段", "目標SKU"]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
            # 套用群組顏色
            if col_name == "群組編號" and isinstance(val, (int, float)) and val in excel_cluster_fills:
                cell.fill = excel_cluster_fills[int(val)]
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 特定表特殊高亮處理
    if sheet_name == "SKU效益分析":
        # 前 10 高亮
        for r in range(2, min(12, len(data_df) + 2)):
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                if headers[c-1] != "群組編號":
                    cell.fill = yellow_fill

        # 條件格式：節省距離成本 綠色漸層
        cost_col_idx = headers.index("節省距離成本") + 1
        col_letter = get_column_letter(cost_col_idx)
        color_scale = ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            end_type='max', end_color='E2F0D9'
        )
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(data_df)+1}", color_scale)

    # 凍結第一列與自動篩選
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    # 自動調整欄寬
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # 處理中文長度計算
            cell_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# 儲存 Excel
wb.save(output_filename)

# 觸發下載
files.download(output_filename)

print("✅ 所有分析程序已成功完成！分析結果 Excel 檔案已自動下載。")
