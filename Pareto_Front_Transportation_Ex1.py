# ============================================================
# 智慧物流教學：時間與成本之多目標物流派車最佳化 Example1
# 作者：國立雲林科技大學電機系 林家仁
# 輸入：Pareto_Setup_File_Ex1
# Time Function + Cost Function + Pareto Front
# 以多目標排程最佳化概念設計
# 讓學員了解Pareto-Front之時間與成本函數最佳解設計
# ============================================================

# ============================================================
# 1. 安裝與載入套件
# ============================================================

!pip -q install openpyxl

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from google.colab import files
from IPython.display import display, Markdown

from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

warnings.filterwarnings("ignore")

# 圖表內容全部使用英文，因此直接使用內建英文字型
plt.rcParams["font.family"] = "DejaVu Sans"
plt.rcParams["axes.unicode_minus"] = False


# ============================================================
# 2. 大地暖色系
# ============================================================

色彩 = {
    "深咖啡": "#665548",
    "中咖啡": "#8B7563",
    "淺咖啡": "#B5A493",

    "米白": "#F8F5F0",
    "淺米色": "#EFE8DE",
    "淡褐色": "#DDD0C3",

    "淺綠": "#E3E9DA",
    "中綠": "#87986A",
    "深綠": "#52633F",

    "淡粉": "#F1DFD9",
    "淺橘": "#F3E1CD",
    "橘棕": "#C6803D",

    "紅色": "#B74435",

    "淺灰": "#E4E1DD",
    "灰色": "#BDBAB6",
    "深灰": "#4D4945",

    "白色": "#FFFFFF"
}


# ============================================================
# 3. 上傳 Excel
# ============================================================

print("請上傳 Excel 檔案，例如：Pareto_Setup_File_Ex1.xlsx")

uploaded = files.upload()

if len(uploaded) == 0:
    raise ValueError("沒有上傳 Excel 檔案，請重新執行程式。")

excel_filename = next(iter(uploaded.keys()))

print(f"\n成功讀取檔案：{excel_filename}")


# ============================================================
# 4. 預設模型參數
#
# 若 Excel 找不到對應參數，使用以下預設值
# ============================================================

參數 = {
    "待配送貨量": 600,
    "每車每趟容量": 50,
    "每日基本趟次": 2,
    "固定理貨時間": 2,
    "每日正常工時": 8,

    "每件處理成本": 15,
    "每車固定成本": 1200,
    "每趟運行成本": 800,
    "駕駛每小時成本": 250,
    "加開趟次成本倍率": 1.5,

    "最少車輛數": 1,
    "最多車輛數": 8,
    "最少加開趟次": 0,
    "最多加開趟次": 3
}


# ============================================================
# 5. 讀取 Excel 所有工作表
# ============================================================

try:

    excel_sheets = pd.read_excel(
        excel_filename,
        sheet_name=None,
        header=None,
        engine="openpyxl"
    )

except Exception as error:

    raise RuntimeError(
        f"Excel 讀取失敗：{error}"
    )


print("\n讀取到的工作表：")

for sheet_name in excel_sheets:
    print(f"－{sheet_name}")


# ============================================================
# 6. Excel 欄位名稱對照
# ============================================================

欄位關鍵字 = {

    "待配送貨量": [
        "待配送貨量",
        "配送需求量",
        "配送需求",
        "配送件數",
        "需求量",
        "總件數",
        "貨物量"
    ],

    "每車每趟容量": [
        "每車每趟容量",
        "每趟容量",
        "單趟容量",
        "車輛容量",
        "每車容量"
    ],

    "每日基本趟次": [
        "每日基本趟次",
        "每車每日基本趟次",
        "每車基本趟次",
        "基本趟次"
    ],

    "固定理貨時間": [
        "固定理貨時間",
        "理貨時間",
        "固定作業時間",
        "準備時間"
    ],

    "每日正常工時": [
        "每日正常工時",
        "每日工時",
        "正常工時",
        "工作時數"
    ],

    "每件處理成本": [
        "每件處理成本",
        "單件處理成本",
        "貨物處理成本",
        "每件成本"
    ],

    "每車固定成本": [
        "每車固定成本",
        "每台車固定成本",
        "車輛固定成本"
    ],

    "每趟運行成本": [
        "每趟運行成本",
        "單趟運行成本",
        "每趟成本",
        "車趟成本"
    ],

    "駕駛每小時成本": [
        "駕駛每小時成本",
        "駕駛時薪",
        "每小時人力成本",
        "駕駛人力成本"
    ],

    "加開趟次成本倍率": [
        "加開趟次成本倍率",
        "加開倍率",
        "加班倍率",
        "額外趟次倍率"
    ],

    "最少車輛數": [
        "最少車輛數",
        "最低車輛數",
        "車輛數下限"
    ],

    "最多車輛數": [
        "最多車輛數",
        "最高車輛數",
        "車輛數上限"
    ],

    "最少加開趟次": [
        "最少加開趟次",
        "加開趟次下限"
    ],

    "最多加開趟次": [
        "最多加開趟次",
        "加開趟次上限"
    ]
}


# ============================================================
# 7. Excel 資料清理函數
# ============================================================

def 清理文字(value):
    """移除空白、換行及冒號，方便比對參數名稱。"""

    if pd.isna(value):
        return ""

    return (
        str(value)
        .strip()
        .replace(" ", "")
        .replace("\n", "")
        .replace("\r", "")
        .replace("：", "")
        .replace(":", "")
    )


def 轉換數值(value):
    """將 Excel 儲存格內容轉換為數值。"""

    if pd.isna(value):
        return None

    if isinstance(
        value,
        (int, float, np.integer, np.floating)
    ):
        return float(value)

    text = str(value).strip()

    移除單位 = [
        ",",
        "元",
        "件",
        "台",
        "趟",
        "小時",
        "時",
        "天",
        "%"
    ]

    for item in 移除單位:
        text = text.replace(item, "")

    try:
        return float(text)

    except (ValueError, TypeError):
        return None


def 搜尋工作表參數(df, keyword_list):
    """
    搜尋工作表中的參數名稱。
    找到參數名稱後，優先讀取右側儲存格，
    若右側無數值，再讀取下方儲存格。
    """

    if df is None or df.empty:
        return None

    row_count, col_count = df.shape

    for row_index in range(row_count):

        for col_index in range(col_count):

            cell_text = 清理文字(
                df.iloc[row_index, col_index]
            )

            if cell_text == "":
                continue

            for keyword in keyword_list:

                keyword_text = 清理文字(keyword)

                matched = (
                    cell_text == keyword_text
                    or keyword_text in cell_text
                    or cell_text in keyword_text
                )

                if not matched:
                    continue

                # 搜尋右側三格
                for offset in range(1, 4):

                    right_col = col_index + offset

                    if right_col < col_count:

                        value = 轉換數值(
                            df.iloc[row_index, right_col]
                        )

                        if value is not None:
                            return value

                # 搜尋下方三格
                for offset in range(1, 4):

                    lower_row = row_index + offset

                    if lower_row < row_count:

                        value = 轉換數值(
                            df.iloc[lower_row, col_index]
                        )

                        if value is not None:
                            return value

    return None


# ============================================================
# 8. 從 Excel 更新模型參數
# ============================================================

參數來源 = {}

for parameter_name, keyword_list in 欄位關鍵字.items():

    found_value = None
    found_sheet = None

    for sheet_name, sheet_df in excel_sheets.items():

        found_value = 搜尋工作表參數(
            sheet_df,
            keyword_list
        )

        if found_value is not None:

            found_sheet = sheet_name
            break

    if found_value is not None:

        參數[parameter_name] = found_value
        參數來源[parameter_name] = found_sheet

    else:

        參數來源[parameter_name] = "程式預設值"


# ============================================================
# 9. 整數參數轉換
# ============================================================

整數參數名稱 = [
    "待配送貨量",
    "每車每趟容量",
    "每日基本趟次",
    "最少車輛數",
    "最多車輛數",
    "最少加開趟次",
    "最多加開趟次"
]

for parameter_name in 整數參數名稱:

    參數[parameter_name] = int(
        round(參數[parameter_name])
    )


# ============================================================
# 10. 參數合理性檢查
# ============================================================

if 參數["待配送貨量"] <= 0:
    raise ValueError("待配送貨量必須大於 0。")

if 參數["每車每趟容量"] <= 0:
    raise ValueError("每車每趟容量必須大於 0。")

if 參數["每日基本趟次"] <= 0:
    raise ValueError("每日基本趟次必須大於 0。")

if 參數["每日正常工時"] <= 0:
    raise ValueError("每日正常工時必須大於 0。")

if 參數["固定理貨時間"] < 0:
    raise ValueError("固定理貨時間不可小於 0。")

if 參數["最少車輛數"] <= 0:
    raise ValueError("最少車輛數必須大於 0。")

if 參數["最多車輛數"] < 參數["最少車輛數"]:
    raise ValueError("最多車輛數不可小於最少車輛數。")

if 參數["最少加開趟次"] < 0:
    raise ValueError("最少加開趟次不可小於 0。")

if 參數["最多加開趟次"] < 參數["最少加開趟次"]:
    raise ValueError("最多加開趟次不可小於最少加開趟次。")

if 參數["加開趟次成本倍率"] < 1:
    raise ValueError("加開趟次成本倍率不可小於 1。")


# ============================================================
# 11. 建立模型參數表
# ============================================================

參數表 = pd.DataFrame({
    "參數名稱": list(參數.keys()),
    "設定值": list(參數.values()),
    "資料來源": [
        參數來源[name]
        for name in 參數.keys()
    ]
})


# ============================================================
# 12. 取得模型變數
# ============================================================

Q = 參數["待配送貨量"]

C = 參數["每車每趟容量"]

b = 參數["每日基本趟次"]

t_s = 參數["固定理貨時間"]

H = 參數["每日正常工時"]

c_h = 參數["每件處理成本"]

c_v = 參數["每車固定成本"]

c_t = 參數["每趟運行成本"]

driver_hourly_cost = 參數["駕駛每小時成本"]

overtime_multiplier = 參數[
    "加開趟次成本倍率"
]

n_min = 參數["最少車輛數"]

n_max = 參數["最多車輛數"]

e_min = 參數["最少加開趟次"]

e_max = 參數["最多加開趟次"]


# 每位駕駛每日人力成本
c_d = driver_hourly_cost * H

# 加開趟次附加成本
c_o = c_t * (
    overtime_multiplier - 1
)


# ============================================================
# 13. 配送時間函數
#
# T(n,e) = ts/H + Q/[nC(b+e)]
# ============================================================

def 配送時間函數(n, e):

    if n <= 0:
        raise ValueError("派遣車輛數必須大於 0。")

    if b + e <= 0:
        raise ValueError("每日總趟次必須大於 0。")

    固定理貨天數 = t_s / H

    每日配送能力 = (
        n
        * C
        * (b + e)
    )

    實際配送天數 = (
        Q / 每日配送能力
    )

    配送完成時間 = (
        固定理貨天數
        + 實際配送天數
    )

    return 配送完成時間


# ============================================================
# 14. 配送成本函數
#
# C(n,e) =
# Qch
# + ncv
# + ct × n × (b+e) × T
# + cd × n × T
# + co × n × e × T
# ============================================================

def 配送成本函數(n, e, completion_time):

    貨物處理成本 = (
        Q * c_h
    )

    車輛固定成本 = (
        n * c_v
    )

    車趟運行成本 = (
        c_t
        * n
        * (b + e)
        * completion_time
    )

    駕駛人力成本 = (
        c_d
        * n
        * completion_time
    )

    加開趟次成本 = (
        c_o
        * n
        * e
        * completion_time
    )

    總配送成本 = (
        貨物處理成本
        + 車輛固定成本
        + 車趟運行成本
        + 駕駛人力成本
        + 加開趟次成本
    )

    return {
        "貨物處理成本": 貨物處理成本,
        "車輛固定成本": 車輛固定成本,
        "車趟運行成本": 車趟運行成本,
        "駕駛人力成本": 駕駛人力成本,
        "加開趟次成本": 加開趟次成本,
        "總配送成本": 總配送成本
    }


# ============================================================
# 15. 建立全部派車方案
# ============================================================

方案資料 = []

方案序號 = 1

for n in range(n_min, n_max + 1):

    for e in range(e_min, e_max + 1):

        completion_time = 配送時間函數(
            n,
            e
        )

        cost_result = 配送成本函數(
            n,
            e,
            completion_time
        )

        每車每日總趟次 = (
            b + e
        )

        每日配送能力 = (
            n
            * C
            * 每車每日總趟次
        )

        方案資料.append({

            "方案編號":
                f"P{方案序號:02d}",

            "派遣車輛數":
                n,

            "每車每日加開趟次":
                e,

            "每車每日總趟次":
                每車每日總趟次,

            "每日配送能力":
                每日配送能力,

            "配送完成時間":
                completion_time,

            "貨物處理成本":
                cost_result["貨物處理成本"],

            "車輛固定成本":
                cost_result["車輛固定成本"],

            "車趟運行成本":
                cost_result["車趟運行成本"],

            "駕駛人力成本":
                cost_result["駕駛人力成本"],

            "加開趟次成本":
                cost_result["加開趟次成本"],

            "總配送成本":
                cost_result["總配送成本"]
        })

        方案序號 += 1


結果表 = pd.DataFrame(
    方案資料
)


# ============================================================
# 16. 判斷 Pareto Front
#
# 若存在另一方案：
# 1. 配送時間不高於目前方案
# 2. 總成本不高於目前方案
# 3. 至少一項嚴格較佳
#
# 則目前方案為被支配解
# ============================================================

def 判斷ParetoFront(df):

    objective_values = df[
        [
            "配送完成時間",
            "總配送成本"
        ]
    ].to_numpy(dtype=float)

    pareto_mask = np.ones(
        len(objective_values),
        dtype=bool
    )

    for i, current_value in enumerate(
        objective_values
    ):

        其他方案不劣 = np.all(
            objective_values <= current_value,
            axis=1
        )

        其他方案至少一項較佳 = np.any(
            objective_values < current_value,
            axis=1
        )

        dominated = np.any(
            其他方案不劣
            &
            其他方案至少一項較佳
        )

        pareto_mask[i] = not dominated

    return pareto_mask


結果表["是否為Pareto解"] = 判斷ParetoFront(
    結果表
)

結果表["方案類型"] = np.where(
    結果表["是否為Pareto解"],
    "非支配解",
    "被支配解"
)


# ============================================================
# 17. 建立非支配解與被支配解表
# ============================================================

非支配解表 = (
    結果表[
        結果表["是否為Pareto解"] == True
    ]
    .copy()
    .sort_values(
        by=[
            "配送完成時間",
            "總配送成本"
        ],
        ascending=[
            True,
            True
        ]
    )
    .reset_index(drop=True)
)


被支配解表 = (
    結果表[
        結果表["是否為Pareto解"] == False
    ]
    .copy()
    .sort_values(
        by=[
            "配送完成時間",
            "總配送成本"
        ],
        ascending=[
            True,
            True
        ]
    )
    .reset_index(drop=True)
)


if 非支配解表.empty:
    raise RuntimeError(
        "未找到 Pareto 非支配解，請檢查模型參數。"
    )


# 後續計算使用 Pareto表 名稱
Pareto表 = 非支配解表.copy()


# ============================================================
# 18. 取得最低成本與最短時間方案
# ============================================================

最低成本方案 = 結果表.loc[
    結果表["總配送成本"].idxmin()
].copy()


最短時間方案 = 結果表.loc[
    結果表["配送完成時間"].idxmin()
].copy()


# ============================================================
# 19. Min-Max 正規化函數
# ============================================================

def min_max_normalize(series):

    minimum = float(series.min())

    maximum = float(series.max())

    if np.isclose(
        maximum,
        minimum
    ):

        return pd.Series(
            np.zeros(len(series)),
            index=series.index
        )

    return (
        (series - minimum)
        /
        (maximum - minimum)
    )


# ============================================================
# 20. Pareto 解正規化
# ============================================================

Pareto表["正規化時間"] = (
    min_max_normalize(
        Pareto表["配送完成時間"]
    )
)

Pareto表["正規化成本"] = (
    min_max_normalize(
        Pareto表["總配送成本"]
    )
)


# ============================================================
# 21. 建議折衷方案
#
# 使用正規化後與理想點 (0,0) 的歐氏距離
# ============================================================

Pareto表["理想點距離"] = np.sqrt(
    Pareto表["正規化時間"] ** 2
    +
    Pareto表["正規化成本"] ** 2
)

建議折衷方案 = Pareto表.loc[
    Pareto表["理想點距離"].idxmin()
].copy()


# ============================================================
# 22. 效益分析
#
# 以最低成本方案作為比較基準
# ============================================================

基準時間 = float(
    最低成本方案["配送完成時間"]
)

基準成本 = float(
    最低成本方案["總配送成本"]
)


Pareto表["節省時間"] = (
    基準時間
    - Pareto表["配送完成時間"]
)


if np.isclose(基準時間, 0):

    Pareto表["時間改善率"] = 0.0

else:

    Pareto表["時間改善率"] = (
        Pareto表["節省時間"]
        /
        基準時間
        * 100
    )


Pareto表["成本差額"] = (
    Pareto表["總配送成本"]
    - 基準成本
)


if np.isclose(基準成本, 0):

    Pareto表["成本增加率"] = 0.0

else:

    Pareto表["成本增加率"] = (
        Pareto表["成本差額"]
        /
        基準成本
        * 100
    )


# ============================================================
# 23. 全部方案顯示表
# ============================================================

方案顯示欄位 = [
    "方案編號",
    "派遣車輛數",
    "每車每日加開趟次",
    "每車每日總趟次",
    "每日配送能力",
    "配送完成時間",
    "總配送成本",
    "方案類型"
]


全部方案顯示表 = (
    結果表[
        方案顯示欄位
    ]
    .copy()
    .reset_index(drop=True)
)


非支配解顯示表 = (
    非支配解表[
        方案顯示欄位
    ]
    .copy()
    .reset_index(drop=True)
)


被支配解顯示表 = (
    被支配解表[
        方案顯示欄位
    ]
    .copy()
    .reset_index(drop=True)
)


# ============================================================
# 24. 支配關係統計表
# ============================================================

總方案數量 = len(結果表)

非支配解數量 = len(非支配解表)

被支配解數量 = len(被支配解表)


支配關係統計表 = pd.DataFrame({

    "方案分類": [
        "全部候選方案",
        "Pareto 非支配解",
        "被支配解"
    ],

    "方案數量": [
        總方案數量,
        非支配解數量,
        被支配解數量
    ],

    "方案占比": [
        100.0,

        非支配解數量
        / 總方案數量
        * 100,

        被支配解數量
        / 總方案數量
        * 100
    ],

    "說明": [
        "所有車輛數與加開趟次的候選組合",

        "不存在其他方案可同時降低配送時間與總配送成本",

        "至少存在一個方案在時間與成本上皆不劣，且其中至少一項更佳"
    ]
})


# ============================================================
# 25. Pareto 時間與成本效益分析表
# ============================================================

效益分析表 = Pareto表[[
    "方案編號",
    "派遣車輛數",
    "每車每日加開趟次",
    "配送完成時間",
    "總配送成本",
    "節省時間",
    "時間改善率",
    "成本差額",
    "成本增加率"
]].copy()


# ============================================================
# 26. 代表性方案比較表
# ============================================================

代表方案表 = pd.DataFrame({

    "方案類型": [
        "最低成本方案",
        "建議折衷方案",
        "最短時間方案"
    ],

    "方案編號": [
        最低成本方案["方案編號"],
        建議折衷方案["方案編號"],
        最短時間方案["方案編號"]
    ],

    "派遣車輛數": [
        int(
            最低成本方案["派遣車輛數"]
        ),

        int(
            建議折衷方案["派遣車輛數"]
        ),

        int(
            最短時間方案["派遣車輛數"]
        )
    ],

    "每車每日加開趟次": [
        int(
            最低成本方案["每車每日加開趟次"]
        ),

        int(
            建議折衷方案["每車每日加開趟次"]
        ),

        int(
            最短時間方案["每車每日加開趟次"]
        )
    ],

    "配送完成時間（天）": [
        float(
            最低成本方案["配送完成時間"]
        ),

        float(
            建議折衷方案["配送完成時間"]
        ),

        float(
            最短時間方案["配送完成時間"]
        )
    ],

    "總配送成本（元）": [
        float(
            最低成本方案["總配送成本"]
        ),

        float(
            建議折衷方案["總配送成本"]
        ),

        float(
            最短時間方案["總配送成本"]
        )
    ]
})


# ============================================================
# 27. 效益分析總表
# ============================================================

折衷節省時間 = (
    float(
        最低成本方案["配送完成時間"]
    )
    -
    float(
        建議折衷方案["配送完成時間"]
    )
)


折衷時間改善率 = (
    折衷節省時間
    /
    float(
        最低成本方案["配送完成時間"]
    )
    * 100
)


折衷相對最短節省成本 = (
    float(
        最短時間方案["總配送成本"]
    )
    -
    float(
        建議折衷方案["總配送成本"]
    )
)


折衷相對最短成本降低率 = (
    折衷相對最短節省成本
    /
    float(
        最短時間方案["總配送成本"]
    )
    * 100
)


效益分析總表 = pd.DataFrame({

    "項目": [
        "時間效益",
        "成本效益",
        "整體效益"
    ],

    "效益說明": [

        (
            f"相較最低成本方案，建議折衷方案可縮短 "
            f"{折衷節省時間:.2f} 天，"
            f"時間改善率為 "
            f"{折衷時間改善率:.2f}%。"
        ),

        (
            f"相較最短時間方案，建議折衷方案可節省 "
            f"{折衷相對最短節省成本:,.0f} 元，"
            f"成本降低率為 "
            f"{折衷相對最短成本降低率:.2f}%。"
        ),

        (
            "建議折衷方案位於 Pareto Front，"
            "可在配送效率與總配送成本之間取得平衡。"
        )
    ]
})


# ============================================================
# 28. 策略建議表
# ============================================================

策略建議表 = pd.DataFrame({

    "策略類型": [
        "成本導向策略",
        "效率導向策略",
        "折衷平衡策略"
    ],

    "適用情境": [
        "貨量穩定、時效要求較低、重視成本控制",

        "高時效訂單、尖峰貨量、客戶交期敏感",

        "一般營運規劃、同時考量配送時間與總成本"
    ],

    "建議方案": [
        最低成本方案["方案編號"],
        最短時間方案["方案編號"],
        建議折衷方案["方案編號"]
    ],

    "策略建議": [

        (
            f'派遣 '
            f'{int(最低成本方案["派遣車輛數"])} 台車，'
            f'每車每日加開 '
            f'{int(最低成本方案["每車每日加開趟次"])} 趟；'
            f'配送時間約 '
            f'{最低成本方案["配送完成時間"]:.2f} 天，'
            f'總成本約 '
            f'{最低成本方案["總配送成本"]:,.0f} 元。'
        ),

        (
            f'派遣 '
            f'{int(最短時間方案["派遣車輛數"])} 台車，'
            f'每車每日加開 '
            f'{int(最短時間方案["每車每日加開趟次"])} 趟；'
            f'配送時間約 '
            f'{最短時間方案["配送完成時間"]:.2f} 天，'
            f'適合高時效與尖峰配送需求。'
        ),

        (
            f'派遣 '
            f'{int(建議折衷方案["派遣車輛數"])} 台車，'
            f'每車每日加開 '
            f'{int(建議折衷方案["每車每日加開趟次"])} 趟；'
            f'配送時間約 '
            f'{建議折衷方案["配送完成時間"]:.2f} 天，'
            f'總成本約 '
            f'{建議折衷方案["總配送成本"]:,.0f} 元，'
            f'可兼顧配送效率與營運成本。'
        )
    ]
})


# ============================================================
# 29. HTML 中文表格共同樣式
#
# 使用 Colab 網頁顯示中文
# 不依賴 Matplotlib 中文字型
# ============================================================

共同表格樣式 = [

    {
        "selector": "table",

        "props": [
            ("border-collapse", "collapse"),
            (
                "font-family",
                "Arial, Microsoft JhengHei, sans-serif"
            ),
            ("font-size", "14px"),
            ("width", "100%"),
            ("background-color", 色彩["白色"])
        ]
    },

    {
        "selector": "thead th",

        "props": [
            ("background-color", 色彩["深咖啡"]),
            ("color", 色彩["白色"]),
            ("font-weight", "bold"),
            ("text-align", "center"),
            ("padding", "10px"),
            (
                "border",
                f'1px solid {色彩["淡褐色"]}'
            ),
            ("white-space", "normal")
        ]
    },

    {
        "selector": "tbody td",

        "props": [
            ("text-align", "center"),
            ("padding", "9px"),
            (
                "border",
                f'1px solid {色彩["淡褐色"]}'
            ),
            ("white-space", "normal")
        ]
    },

    {
        "selector": "tbody tr:hover",

        "props": [
            ("filter", "brightness(0.97)")
        ]
    }
]


# ============================================================
# 30. 表格配色函數
# ============================================================

def 一般交錯列配色(row):

    background = (
        色彩["米白"]
        if row.name % 2 == 0
        else 色彩["淺米色"]
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 全部方案列配色(row):

    if row["方案類型"] == "非支配解":

        background = 色彩["淺綠"]

    else:

        background = (
            色彩["淺灰"]
            if row.name % 2 == 0
            else "#F3F1EE"
        )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 非支配解列配色(row):

    background = (
        色彩["淺綠"]
        if row.name % 2 == 0
        else "#F1F4EC"
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 被支配解列配色(row):

    background = (
        色彩["淺灰"]
        if row.name % 2 == 0
        else "#F4F2EF"
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 支配統計列配色(row):

    color_map = {
        "全部候選方案":
            色彩["淺米色"],

        "Pareto 非支配解":
            色彩["淺綠"],

        "被支配解":
            色彩["淺灰"]
    }

    background = color_map.get(
        row["方案分類"],
        色彩["米白"]
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 代表方案列配色(row):

    color_map = {
        "最低成本方案":
            色彩["淺綠"],

        "建議折衷方案":
            色彩["淡粉"],

        "最短時間方案":
            色彩["淺橘"]
    }

    background = color_map.get(
        row["方案類型"],
        色彩["米白"]
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 效益總表列配色(row):

    color_map = {
        "時間效益":
            色彩["淺綠"],

        "成本效益":
            色彩["淺橘"],

        "整體效益":
            色彩["淡粉"]
    }

    background = color_map.get(
        row["項目"],
        色彩["米白"]
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


def 策略建議列配色(row):

    color_map = {
        "成本導向策略":
            色彩["淺綠"],

        "效率導向策略":
            色彩["淺橘"],

        "折衷平衡策略":
            色彩["淡粉"]
    }

    background = color_map.get(
        row["策略類型"],
        色彩["米白"]
    )

    return [
        f"background-color: {background};"
        for _ in row
    ]


# ============================================================
# 31. 資料表格式設定
# ============================================================

方案表格式 = {
    "派遣車輛數":
        "{:.0f} 台",

    "每車每日加開趟次":
        "{:.0f} 趟",

    "每車每日總趟次":
        "{:.0f} 趟",

    "每日配送能力":
        "{:,.0f} 件／日",

    "配送完成時間":
        "{:.2f} 天",

    "總配送成本":
        "{:,.0f} 元"
}


效益分析格式 = {
    "派遣車輛數":
        "{:.0f} 台",

    "每車每日加開趟次":
        "{:.0f} 趟",

    "配送完成時間":
        "{:.2f} 天",

    "總配送成本":
        "{:,.0f} 元",

    "節省時間":
        "{:.2f} 天",

    "時間改善率":
        "{:.2f}%",

    "成本差額":
        "{:+,.0f} 元",

    "成本增加率":
        "{:+.2f}%"
}


代表方案格式 = {
    "派遣車輛數":
        "{:.0f} 台",

    "每車每日加開趟次":
        "{:.0f} 趟",

    "配送完成時間（天）":
        "{:.2f}",

    "總配送成本（元）":
        "{:,.0f}"
}


# ============================================================
# 32. 顯示模型輸入參數表
# ============================================================

display(
    Markdown(
        "## 1．模型輸入參數"
    )
)

display(
    參數表.style
    .apply(
        一般交錯列配色,
        axis=1
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 33. 顯示全部候選方案表
# ============================================================

display(
    Markdown(
        f"## 2．全部候選方案表（共 {len(全部方案顯示表)} 種）"
    )
)

display(
    全部方案顯示表.style
    .format(
        方案表格式
    )
    .apply(
        全部方案列配色,
        axis=1
    )
    .set_properties(
        subset=[
            "方案編號",
            "配送完成時間",
            "總配送成本"
        ],
        **{
            "font-weight": "bold"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 34. 顯示 Pareto 非支配解表
# ============================================================

display(
    Markdown(
        f"## 3．Pareto 非支配解表（共 {len(非支配解顯示表)} 種）"
    )
)

display(
    非支配解顯示表.style
    .format(
        方案表格式
    )
    .apply(
        非支配解列配色,
        axis=1
    )
    .set_properties(
        subset=["方案編號"],
        **{
            "font-weight": "bold",
            "color": 色彩["深綠"]
        }
    )
    .set_properties(
        subset=[
            "配送完成時間",
            "總配送成本"
        ],
        **{
            "font-weight": "bold"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 35. 顯示被支配解表
# ============================================================

display(
    Markdown(
        f"## 4．被支配解表（共 {len(被支配解顯示表)} 種）"
    )
)

display(
    被支配解顯示表.style
    .format(
        方案表格式
    )
    .apply(
        被支配解列配色,
        axis=1
    )
    .set_properties(
        subset=["方案編號"],
        **{
            "font-weight": "bold",
            "color": 色彩["深灰"]
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 36. 顯示支配關係統計表
# ============================================================

display(
    Markdown(
        "## 5．支配關係統計"
    )
)

display(
    支配關係統計表.style
    .format({
        "方案數量":
            "{:.0f} 種",

        "方案占比":
            "{:.2f}%"
    })
    .apply(
        支配統計列配色,
        axis=1
    )
    .set_properties(
        subset=["說明"],
        **{
            "text-align": "left",
            "white-space": "normal"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 37. 顯示 Pareto 時間與成本效益分析表
# ============================================================

display(
    Markdown(
        "## 6．Pareto 方案－時間與成本效益分析"
    )
)

display(
    效益分析表.style
    .format(
        效益分析格式
    )
    .apply(
        一般交錯列配色,
        axis=1
    )
    .set_properties(
        subset=[
            "配送完成時間",
            "節省時間",
            "時間改善率"
        ],
        **{
            "background-color":
                色彩["淺綠"]
        }
    )
    .set_properties(
        subset=[
            "成本差額",
            "成本增加率"
        ],
        **{
            "background-color":
                色彩["淡粉"]
        }
    )
    .set_properties(
        subset=["方案編號"],
        **{
            "font-weight": "bold"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 38. 顯示代表性方案比較表
# ============================================================

display(
    Markdown(
        "## 7．智慧物流轉運中心－代表性方案比較"
    )
)

display(
    代表方案表.style
    .format(
        代表方案格式
    )
    .apply(
        代表方案列配色,
        axis=1
    )
    .set_properties(
        subset=[
            "方案類型",
            "方案編號"
        ],
        **{
            "font-weight": "bold"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 39. 顯示效益分析總表
# ============================================================

display(
    Markdown(
        "## 8．效益分析總表"
    )
)

display(
    效益分析總表.style
    .apply(
        效益總表列配色,
        axis=1
    )
    .set_properties(
        subset=["項目"],
        **{
            "font-weight": "bold"
        }
    )
    .set_properties(
        subset=["效益說明"],
        **{
            "text-align": "left",
            "white-space": "normal"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 40. 顯示策略建議表
# ============================================================

display(
    Markdown(
        "## 9．物流配送策略建議"
    )
)

display(
    策略建議表.style
    .apply(
        策略建議列配色,
        axis=1
    )
    .set_properties(
        subset=[
            "策略類型",
            "建議方案"
        ],
        **{
            "font-weight": "bold"
        }
    )
    .set_properties(
        subset=[
            "適用情境",
            "策略建議"
        ],
        **{
            "text-align": "left",
            "white-space": "normal"
        }
    )
    .set_table_styles(
        共同表格樣式
    )
    .hide(
        axis="index"
    )
)


# ============================================================
# 41. Pareto Front
#
# 圖內全部使用英文
# 不儲存圖片
# 不下載圖片
# ============================================================

fig, ax = plt.subplots(
    figsize=(13, 7)
)


# Dominated solutions
ax.scatter(
    被支配解表["配送完成時間"],
    被支配解表["總配送成本"],

    s=65,

    color=色彩["灰色"],

    alpha=0.72,

    edgecolors=色彩["白色"],

    linewidths=0.7,

    label="Dominated Solutions"
)


# Pareto solutions
ax.scatter(
    Pareto表["配送完成時間"],
    Pareto表["總配送成本"],

    s=105,

    color=色彩["橘棕"],

    edgecolors=色彩["深咖啡"],

    linewidths=1.0,

    label="Pareto Solutions",

    zorder=3
)


# Pareto front line
ax.plot(
    Pareto表["配送完成時間"],
    Pareto表["總配送成本"],

    color=色彩["深咖啡"],

    linewidth=2.2,

    label="Pareto Front",

    zorder=2
)


# Minimum-cost solution
ax.scatter(
    最低成本方案["配送完成時間"],
    最低成本方案["總配送成本"],

    s=180,

    marker="s",

    color=色彩["中綠"],

    edgecolors=色彩["深綠"],

    linewidths=1.2,

    label="Minimum-Cost Solution",

    zorder=4
)


# Minimum-time solution
ax.scatter(
    最短時間方案["配送完成時間"],
    最短時間方案["總配送成本"],

    s=185,

    marker="D",

    color=色彩["中咖啡"],

    edgecolors=色彩["深咖啡"],

    linewidths=1.2,

    label="Minimum-Time Solution",

    zorder=4
)


# Recommended compromise
ax.scatter(
    建議折衷方案["配送完成時間"],
    建議折衷方案["總配送成本"],

    s=350,

    marker="*",

    color=色彩["紅色"],

    edgecolors="#7E271E",

    linewidths=1.2,

    label="Recommended Compromise",

    zorder=5
)


# Pareto solution labels
for _, row in Pareto表.iterrows():

    ax.annotate(
        row["方案編號"],

        (
            row["配送完成時間"],
            row["總配送成本"]
        ),

        xytext=(5, 6),

        textcoords="offset points",

        fontsize=8,

        color=色彩["深灰"]
    )


# Recommended solution annotation
ax.annotate(

    (
        f'Recommended: '
        f'{建議折衷方案["方案編號"]}\n'

        f'Vehicles: '
        f'{int(建議折衷方案["派遣車輛數"])}\n'

        f'Extra Trips: '
        f'{int(建議折衷方案["每車每日加開趟次"])}\n'

        f'Time: '
        f'{建議折衷方案["配送完成時間"]:.2f} Days\n'

        f'Cost: NTD '
        f'{建議折衷方案["總配送成本"]:,.0f}'
    ),

    (
        建議折衷方案["配送完成時間"],
        建議折衷方案["總配送成本"]
    ),

    xytext=(40, 30),

    textcoords="offset points",

    fontsize=9,

    color=色彩["深灰"],

    bbox={
        "boxstyle": "round,pad=0.5",
        "facecolor": 色彩["白色"],
        "edgecolor": 色彩["中咖啡"],
        "alpha": 0.96
    },

    arrowprops={
        "arrowstyle": "->",
        "color": 色彩["中咖啡"]
    }
)


ax.set_title(
    "Pareto Front: Delivery Time vs. Total Cost",

    fontsize=16,

    fontweight="bold",

    color=色彩["深咖啡"],

    pad=15
)


ax.set_xlabel(
    "Delivery Completion Time (Days)",

    fontsize=12
)


ax.set_ylabel(
    "Total Delivery Cost (NTD)",

    fontsize=12
)


ax.grid(
    alpha=0.22,

    linestyle="--"
)


ax.legend(
    frameon=True,

    facecolor=色彩["白色"],

    edgecolor=色彩["淡褐色"],

    fontsize=9
)


ax.ticklabel_format(
    style="plain",

    axis="y"
)


fig.tight_layout()

plt.show()


# ============================================================
# 42. 匯出 Excel
#
# 只建立 Excel 檔案
# 不自動下載
# ============================================================

output_excel = (
    "物流配送_Pareto完整分析結果.xlsx"
)


with pd.ExcelWriter(
    output_excel,
    engine="openpyxl"
) as writer:

    參數表.to_excel(
        writer,
        sheet_name="模型參數",
        index=False
    )

    結果表.to_excel(
        writer,
        sheet_name="全部方案",
        index=False
    )

    非支配解表.to_excel(
        writer,
        sheet_name="非支配解",
        index=False
    )

    被支配解表.to_excel(
        writer,
        sheet_name="被支配解",
        index=False
    )

    支配關係統計表.to_excel(
        writer,
        sheet_name="支配關係統計",
        index=False
    )

    Pareto表.to_excel(
        writer,
        sheet_name="Pareto效益資料",
        index=False
    )

    效益分析表.to_excel(
        writer,
        sheet_name="效益分析",
        index=False
    )

    代表方案表.to_excel(
        writer,
        sheet_name="代表方案比較",
        index=False
    )

    效益分析總表.to_excel(
        writer,
        sheet_name="效益分析總表",
        index=False
    )

    策略建議表.to_excel(
        writer,
        sheet_name="策略建議",
        index=False
    )


# ============================================================
# 43. Excel 大地暖色系格式
# ============================================================

workbook = load_workbook(
    output_excel
)


# 標題列
header_fill = PatternFill(
    fill_type="solid",
    fgColor="665548"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)


# 一般交錯色
odd_fill = PatternFill(
    fill_type="solid",
    fgColor="F8F5F0"
)

even_fill = PatternFill(
    fill_type="solid",
    fgColor="EFE8DE"
)


# 特殊列色彩
green_fill = PatternFill(
    fill_type="solid",
    fgColor="E3E9DA"
)

pink_fill = PatternFill(
    fill_type="solid",
    fgColor="F1DFD9"
)

orange_fill = PatternFill(
    fill_type="solid",
    fgColor="F3E1CD"
)

gray_fill = PatternFill(
    fill_type="solid",
    fgColor="E4E1DD"
)

light_gray_fill = PatternFill(
    fill_type="solid",
    fgColor="F4F2EF"
)


thin_border = Border(

    left=Side(
        style="thin",
        color="D6C9BD"
    ),

    right=Side(
        style="thin",
        color="D6C9BD"
    ),

    top=Side(
        style="thin",
        color="D6C9BD"
    ),

    bottom=Side(
        style="thin",
        color="D6C9BD"
    )
)


# ============================================================
# 44. 所有 Excel 工作表基本格式
# ============================================================

for worksheet in workbook.worksheets:

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # 標題列
    for cell in worksheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    # 資料列
    for row_index in range(
        2,
        worksheet.max_row + 1
    ):

        row_fill = (
            odd_fill
            if row_index % 2 == 0
            else even_fill
        )

        for col_index in range(
            1,
            worksheet.max_column + 1
        ):

            cell = worksheet.cell(
                row=row_index,
                column=col_index
            )

            cell.fill = row_fill

            cell.border = thin_border

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # 自動欄寬
    for column_cells in worksheet.columns:

        column_letter = (
            column_cells[0].column_letter
        )

        maximum_length = 0

        for cell in column_cells:

            if cell.value is not None:

                cell_length = len(
                    str(cell.value)
                )

                maximum_length = max(
                    maximum_length,
                    cell_length
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            maximum_length + 4,
            45
        )

    # 資料列高度
    for row_index in range(
        2,
        worksheet.max_row + 1
    ):

        worksheet.row_dimensions[
            row_index
        ].height = 24


# ============================================================
# 45. Excel 全部方案表配色
# ============================================================

if "全部方案" in workbook.sheetnames:

    ws = workbook["全部方案"]

    header_names = {
        cell.value: cell.column
        for cell in ws[1]
    }

    type_column = header_names.get(
        "方案類型"
    )

    if type_column is not None:

        for row_index in range(
            2,
            ws.max_row + 1
        ):

            solution_type = ws.cell(
                row=row_index,
                column=type_column
            ).value

            if solution_type == "非支配解":

                row_fill = green_fill

            else:

                row_fill = (
                    gray_fill
                    if row_index % 2 == 0
                    else light_gray_fill
                )

            for col_index in range(
                1,
                ws.max_column + 1
            ):

                ws.cell(
                    row=row_index,
                    column=col_index
                ).fill = row_fill


# ============================================================
# 46. Excel 非支配解工作表配色
# ============================================================

if "非支配解" in workbook.sheetnames:

    ws = workbook["非支配解"]

    for row_index in range(
        2,
        ws.max_row + 1
    ):

        row_fill = (
            green_fill
            if row_index % 2 == 0
            else odd_fill
        )

        for col_index in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                row=row_index,
                column=col_index
            ).fill = row_fill


# ============================================================
# 47. Excel 被支配解工作表配色
# ============================================================

if "被支配解" in workbook.sheetnames:

    ws = workbook["被支配解"]

    for row_index in range(
        2,
        ws.max_row + 1
    ):

        row_fill = (
            gray_fill
            if row_index % 2 == 0
            else light_gray_fill
        )

        for col_index in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                row=row_index,
                column=col_index
            ).fill = row_fill


# ============================================================
# 48. Excel 支配關係統計表配色
# ============================================================

if "支配關係統計" in workbook.sheetnames:

    ws = workbook["支配關係統計"]

    row_fill_list = [
        even_fill,
        green_fill,
        gray_fill
    ]

    for row_index, row_fill in zip(
        range(2, 5),
        row_fill_list
    ):

        for col_index in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                row=row_index,
                column=col_index
            ).fill = row_fill


# ============================================================
# 49. Excel 代表方案比較表配色
# ============================================================

if "代表方案比較" in workbook.sheetnames:

    ws = workbook["代表方案比較"]

    row_fill_list = [
        green_fill,
        pink_fill,
        orange_fill
    ]

    for row_index, row_fill in zip(
        range(2, 5),
        row_fill_list
    ):

        for col_index in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                row=row_index,
                column=col_index
            ).fill = row_fill


# ============================================================
# 50. Excel 效益分析總表配色
# ============================================================

if "效益分析總表" in workbook.sheetnames:

    ws = workbook["效益分析總表"]

    row_fill_list = [
        green_fill,
        orange_fill,
        pink_fill
    ]

    for row_index, row_fill in zip(
        range(2, 5),
        row_fill_list
    ):

        for col_index in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                row=row_index,
                column=col_index
            ).fill = row_fill


# ============================================================
# 51. Excel 策略建議表配色
# ============================================================

if "策略建議" in workbook.sheetnames:

    ws = workbook["策略建議"]

    row_fill_list = [
        green_fill,
        orange_fill,
        pink_fill
    ]

    for row_index, row_fill in zip(
        range(2, 5),
        row_fill_list
    ):

        for col_index in range(
            1,
            ws.max_column + 1
        ):

            ws.cell(
                row=row_index,
                column=col_index
            ).fill = row_fill

    # 策略建議欄加寬
    header_names = {
        cell.value: cell.column_letter
        for cell in ws[1]
    }

    if "適用情境" in header_names:

        ws.column_dimensions[
            header_names["適用情境"]
        ].width = 35

    if "策略建議" in header_names:

        ws.column_dimensions[
            header_names["策略建議"]
        ].width = 55


# ============================================================
# 52. Excel 效益分析欄位格式
# ============================================================

if "效益分析" in workbook.sheetnames:

    ws = workbook["效益分析"]

    header_names = {
        cell.value: cell.column
        for cell in ws[1]
    }

    number_formats = {
        "配送完成時間": "0.00",
        "總配送成本": '#,##0',
        "節省時間": "0.00",
        "時間改善率": '0.00"%"',
        "成本差額": '+#,##0;-#,##0;0',
        "成本增加率": '+0.00"%" ;-0.00"%" ;0.00"%"'
    }

    for header_name, number_format in (
        number_formats.items()
    ):

        column_index = header_names.get(
            header_name
        )

        if column_index is None:
            continue

        for row_index in range(
            2,
            ws.max_row + 1
        ):

            ws.cell(
                row=row_index,
                column=column_index
            ).number_format = number_format


# ============================================================
# 53. 儲存 Excel
# ============================================================

workbook.save(
    output_excel
)


# ============================================================
# 54. 顯示分析摘要
# ============================================================

display(
    Markdown(
        "## 10．分析摘要"
    )
)

print(
    f"全部候選方案："
    f"{總方案數量} 種"
)

print(
    f"Pareto 非支配解："
    f"{非支配解數量} 種"
)

print(
    f"被支配解："
    f"{被支配解數量} 種"
)


print("\n最低成本方案")

print(
    f"方案編號："
    f"{最低成本方案['方案編號']}"
)

print(
    f"派遣車輛數："
    f"{int(最低成本方案['派遣車輛數'])} 台"
)

print(
    f"每車每日加開趟次："
    f"{int(最低成本方案['每車每日加開趟次'])} 趟"
)

print(
    f"配送完成時間："
    f"{最低成本方案['配送完成時間']:.2f} 天"
)

print(
    f"總配送成本："
    f"{最低成本方案['總配送成本']:,.0f} 元"
)


print("\n建議折衷方案")

print(
    f"方案編號："
    f"{建議折衷方案['方案編號']}"
)

print(
    f"派遣車輛數："
    f"{int(建議折衷方案['派遣車輛數'])} 台"
)

print(
    f"每車每日加開趟次："
    f"{int(建議折衷方案['每車每日加開趟次'])} 趟"
)

print(
    f"配送完成時間："
    f"{建議折衷方案['配送完成時間']:.2f} 天"
)

print(
    f"總配送成本："
    f"{建議折衷方案['總配送成本']:,.0f} 元"
)


print("\n最短時間方案")

print(
    f"方案編號："
    f"{最短時間方案['方案編號']}"
)

print(
    f"派遣車輛數："
    f"{int(最短時間方案['派遣車輛數'])} 台"
)

print(
    f"每車每日加開趟次："
    f"{int(最短時間方案['每車每日加開趟次'])} 趟"
)

print(
    f"配送完成時間："
    f"{最短時間方案['配送完成時間']:.2f} 天"
)

print(
    f"總配送成本："
    f"{最短時間方案['總配送成本']:,.0f} 元"
)


print("\n程式執行完成。")

print(
    f"Excel 結果已建立：{output_excel}"
)

print(
    "請在 Colab 左側「檔案」區手動下載 Excel。"
)

print(
    "本程式不會儲存或下載任何圖片。"
)
