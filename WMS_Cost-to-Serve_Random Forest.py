# ============================================================
# 物流業 WMS AI 智慧成本分析實作
# Random Forest - Cost-to-Serve Prediction
# Public Teaching Version | Warm Dashboard Style
# 作者：國立雲林科技大學電機系 林家仁
# ★ 與 Multiple Regression 使用相同資料邏輯
# ============================================================


# ============================================================
# 0. Import Packages
# ============================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from IPython.display import display, HTML
from google.colab import files

from sklearn.model_selection import GroupShuffleSplit
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline

from sklearn.ensemble import RandomForestRegressor

from sklearn.metrics import (
    r2_score,
    mean_absolute_error,
    mean_squared_error,
    mean_absolute_percentage_error
)

from matplotlib.colors import LinearSegmentedColormap

import warnings
warnings.filterwarnings("ignore")


# ============================================================
# 1. Warm Color Palette
# ============================================================

COLORS = {

    "cream": "#FFF8F0",
    "cream2": "#FFF2E5",

    "peach": "#F7C9A9",
    "peach_dark": "#E9A77F",

    "coral": "#E98F7C",
    "terracotta": "#C96F5B",

    "rose": "#D7A3A9",

    "gold": "#D9A15B",

    "sage": "#A8B89A",
    "sage_dark": "#7F9273",

    "brown": "#6D5142",
    "brown_dark": "#4F3A30",

    "gray": "#746C67",

    "white": "#FFFFFF"
}


# ------------------------------------------------------------
# Package Size Colors
# ------------------------------------------------------------

SIZE_COLORS = {

    "S60": "#A8B89A",

    "S90": "#F2C879",

    "S120": "#E9A77F",

    "S150": "#D77F73"
}


# ------------------------------------------------------------
# English labels
# ------------------------------------------------------------

REGION_EN = {

    "北部": "North",

    "中部": "Central",

    "南部": "South",

    "東部": "East",

    "離島": "Offshore"
}


TEMPERATURE_EN = {

    "常溫": "Ambient",

    "低溫": "Chilled"
}


INDUSTRY_EN = {

    "生醫用品": "Biomedical",

    "食品冷鏈": "Cold Chain Food",

    "書籍文具": "Books & Stationery",

    "汽車零件": "Auto Parts",

    "電商零售": "E-commerce",

    "美妝保健": "Beauty & Health",

    "服飾鞋包": "Fashion",

    "3C家電": "Electronics"
}


REGION_COLORS = {

    "North": "#A8B89A",

    "Central": "#F2C879",

    "South": "#E9A77F",

    "East": "#D7A3A9",

    "Offshore": "#C96F5B"
}


TEMP_COLORS = {

    "Ambient": "#E9A77F",

    "Chilled": "#A8B89A"
}


# ------------------------------------------------------------
# Warm Gradient
# ------------------------------------------------------------

warm_cmap = LinearSegmentedColormap.from_list(

    "warm_gradient",

    [
        "#FFF8F0",
        "#F7C9A9",
        "#E98F7C"
    ]
)


# ============================================================
# 2. Matplotlib Style
# ============================================================

plt.rcParams["figure.figsize"] = (10, 6)

plt.rcParams["axes.titleweight"] = "bold"

plt.rcParams["axes.titlesize"] = 15

plt.rcParams["axes.labelsize"] = 11

plt.rcParams["axes.edgecolor"] = COLORS["brown"]

plt.rcParams["axes.labelcolor"] = COLORS["brown_dark"]

plt.rcParams["xtick.color"] = COLORS["gray"]

plt.rcParams["ytick.color"] = COLORS["gray"]

plt.rcParams["font.family"] = "DejaVu Sans"


# ============================================================
# 3. Dashboard Functions
# ============================================================

def dashboard_title(title, subtitle=""):

    html = f"""
    <div style="
        background:linear-gradient(135deg,#FFF8F0,#F7C9A9);
        border-left:8px solid #C96F5B;
        padding:22px 28px;
        margin:16px 0 22px 0;
        border-radius:14px;
        box-shadow:0 3px 10px rgba(100,70,50,0.10);
    ">

        <div style="
            font-size:26px;
            font-weight:800;
            color:#4F3A30;
            margin-bottom:6px;
        ">
            {title}
        </div>

        <div style="
            font-size:15px;
            color:#746C67;
            line-height:1.8;
        ">
            {subtitle}
        </div>

    </div>
    """

    display(HTML(html))


def section_title(number, title, description=""):

    html = f"""
    <div style="
        margin-top:26px;
        margin-bottom:14px;
    ">

        <div style="
            display:inline-block;
            background:#C96F5B;
            color:white;
            font-size:14px;
            font-weight:bold;
            padding:6px 12px;
            border-radius:18px;
            margin-bottom:8px;
        ">
            STEP {number}
        </div>

        <div style="
            font-size:21px;
            font-weight:800;
            color:#4F3A30;
        ">
            {title}
        </div>

        <div style="
            font-size:14px;
            color:#746C67;
            margin-top:5px;
            line-height:1.7;
        ">
            {description}
        </div>

    </div>
    """

    display(HTML(html))


def info_box(text):

    html = f"""
    <div style="
        background:#FFF8F0;
        border:1px solid #F0D7C4;
        padding:14px 18px;
        border-radius:10px;
        color:#6D5142;
        line-height:1.8;
        margin:10px 0 16px 0;
    ">
        💡 {text}
    </div>
    """

    display(HTML(html))


def metric_cards(metrics):

    cards = ""

    card_colors = [
        "#F7C9A9",
        "#F2C879",
        "#A8B89A",
        "#D7A3A9"
    ]

    for i, (name, value, note) in enumerate(metrics):

        color = card_colors[
            i % len(card_colors)
        ]

        cards += f"""
        <div style="
            flex:1;
            min-width:160px;
            background:white;
            border-top:6px solid {color};
            border-radius:12px;
            padding:16px;
            margin:6px;
            box-shadow:0 3px 10px rgba(80,60,50,0.08);
        ">

            <div style="
                color:#746C67;
                font-size:13px;
            ">
                {name}
            </div>

            <div style="
                color:#4F3A30;
                font-size:28px;
                font-weight:800;
                margin:6px 0;
            ">
                {value}
            </div>

            <div style="
                color:#8A7F79;
                font-size:12px;
            ">
                {note}
            </div>

        </div>
        """

    html = f"""
    <div style="
        display:flex;
        flex-wrap:wrap;
        margin:12px -6px 18px -6px;
    ">
        {cards}
    </div>
    """

    display(HTML(html))


def warm_style(
    df,
    formats=None,
    gradient_cols=None,
    hide_index=True
):

    styler = (

        df.style

        .set_table_styles([

            {
                "selector": "thead th",

                "props": [

                    ("background-color", COLORS["terracotta"]),

                    ("color", "white"),

                    ("font-weight", "bold"),

                    ("text-align", "center"),

                    ("padding", "10px")
                ]
            },

            {
                "selector": "tbody td",

                "props": [

                    ("padding", "8px"),

                    ("border-bottom", "1px solid #F1E4DA"),

                    ("color", COLORS["brown_dark"])
                ]
            },

            {
                "selector": "tbody tr:nth-child(even)",

                "props": [

                    ("background-color", COLORS["cream"])
                ]
            },

            {
                "selector": "tbody tr:hover",

                "props": [

                    ("background-color", COLORS["cream2"])
                ]
            },

            {
                "selector": "table",

                "props": [

                    ("border-collapse", "collapse"),

                    ("font-size", "13px")
                ]
            }

        ])
    )


    if formats:

        styler = styler.format(
            formats
        )


    if gradient_cols:

        valid_cols = [

            col

            for col in gradient_cols

            if col in df.columns
        ]


        if valid_cols:

            styler = styler.background_gradient(

                cmap=warm_cmap,

                subset=valid_cols
            )


    if hide_index:

        try:

            styler = styler.hide(
                axis="index"
            )

        except:

            pass


    return styler


# ============================================================
# 4. Translate Feature Names for Charts
# ============================================================

def feature_name_to_english(name):

    text = str(name)


    # Package Size

    text = text.replace(

        "Package_Size_",

        "Package Size: "
    )


    # Temperature

    text = text.replace(

        "Temperature_低溫",

        "Temperature: Chilled"
    )

    text = text.replace(

        "Temperature_常溫",

        "Temperature: Ambient"
    )


    # Region

    for zh, en in REGION_EN.items():

        text = text.replace(

            f"Destination_Region_{zh}",

            f"Region: {en}"
        )


    # Industry

    for zh, en in INDUSTRY_EN.items():

        text = text.replace(

            f"Industry_{zh}",

            f"Industry: {en}"
        )


    # Customer Type

    text = text.replace(

        "Customer_Type_",

        "Customer Type: "
    )


    # Numeric Feature

    replacements = {

        "Monthly_Target_Packages":
            "Monthly Package Volume",

        "Weight_kg":
            "Package Weight",

        "Distance_km":
            "Delivery Distance",

        "Remote_Flag":
            "Remote Delivery",

        "COD_Flag":
            "COD Service",

        "COD_Amount":
            "COD Amount",

        "Insurance_Flag":
            "Insurance Service",

        "Declared_Value":
            "Declared Value",

        "Redelivery_Flag":
            "Redelivery",

        "SameDay_Flag":
            "Same-Day Delivery",

        "StorePickup_Flag":
            "Store Pickup",

        "Pickup_Density":
            "Pickup Density",

        "Customer_Cold_Ratio":
            "Customer Chilled Ratio",

        "Customer_Remote_Ratio":
            "Customer Remote Ratio",

        "Customer_Redelivery_Rate":
            "Customer Redelivery Rate"
    }


    if text in replacements:

        text = replacements[text]


    return text


# ============================================================
# 5. Main Dashboard
# ============================================================

dashboard_title(

    "物流業 WMS｜Random Forest 成本預測",

    """
    本實作使用與 Multiple Regression 相同的
    多客戶、多包裹 WMS 資料，
    建立 Random Forest Cost-to-Serve 模型，
    進一步學習物流成本中的非線性、
    門檻效應與多條件交互作用。
    """
)


# ============================================================
# STEP 1 - Upload Excel
# ============================================================

section_title(

    1,

    "上傳物流業 WMS 教學資料",

    "請上傳與 Multiple Regression 相同的 WMS Excel 資料。"
)


uploaded = files.upload()


filename = list(
    uploaded.keys()
)[0]


info_box(

    f"已成功載入物流業 WMS 資料：<b>{filename}</b>"
)


# ============================================================
# STEP 2 - Read Data
# ============================================================

section_title(

    2,

    "讀取 WMS Package Data",

    """
    本分析仍然採用一列一包裹的 Package-Level Data，
    確保 Random Forest 與 Multiple Regression
    使用完全相同的原始資料。
    """
)


df = pd.read_excel(

    filename,

    sheet_name="WMS_Package_Data"
)


df = df.loc[
    :,
    ~df.columns.astype(str).str.startswith("Unnamed")
]


if "index" in df.columns:

    df = df.drop(
        columns=["index"]
    )


metric_cards([

    (
        "WMS Packages",

        f"{len(df):,}",

        "每列 = 1 件包裹"
    ),

    (
        "Customers",

        f"{df['Customer_ID'].nunique():,}",

        "多客戶資料"
    ),

    (
        "Columns",

        f"{len(df.columns)}",

        "WMS Features"
    ),

    (
        "Target",

        "Cost",

        "Actual Operational Cost"
    )
])


# ============================================================
# STEP 3 - Package Size
# ============================================================

section_title(

    3,

    "觀察 S60 / S90 / S120 / S150",

    """
    先確認不同材積級距的資料分布，
    後續將分析 Random Forest
    是否能改善大型包裹的成本預測誤差。
    """
)


size_counts = (

    df["Package_Size"]

    .value_counts()

    .reindex([

        "S60",

        "S90",

        "S120",

        "S150"
    ])

    .fillna(0)

    .astype(int)
)


size_percent = (

    size_counts

    / size_counts.sum()

    * 100
)


size_table = pd.DataFrame({

    "Package Size":
        size_counts.index,

    "Package Count":
        size_counts.values,

    "Percentage (%)":
        size_percent.values
})


display(

    warm_style(

        size_table,

        formats={

            "Package Count":
                "{:,.0f}",

            "Percentage (%)":
                "{:.2f}%"
        }
    )
)


# ============================================================
# Chart 1 - Package Size Distribution
# ============================================================

fig, ax = plt.subplots(
    figsize=(9, 5.5)
)


bars = ax.bar(

    size_counts.index,

    size_counts.values,

    color=[

        SIZE_COLORS[x]

        for x in size_counts.index
    ],

    edgecolor="white",

    linewidth=1.5
)


ax.set_title(
    "Package Size Distribution",
    color=COLORS["brown_dark"],
    pad=15
)


ax.set_xlabel(
    "Package Size"
)


ax.set_ylabel(
    "Number of Packages"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


for bar, value in zip(
    bars,
    size_counts.values
):

    ax.text(

        bar.get_x()
        + bar.get_width()/2,

        value
        + max(size_counts.values)*0.015,

        f"{value:,}",

        ha="center",

        fontweight="bold",

        color=COLORS["brown"]
    )


plt.tight_layout()

plt.show()


# ============================================================
# STEP 4 - Define X and y
# ============================================================

section_title(

    4,

    "建立 Random Forest 的 X 與 y",

    """
    與 Multiple Regression 保持完全一致。
    X 為物流特徵，y 為 Actual Operational Cost。
    """
)


numeric_features = [

    "Monthly_Target_Packages",

    "Weight_kg",

    "Distance_km",

    "Remote_Flag",

    "COD_Flag",

    "COD_Amount",

    "Insurance_Flag",

    "Declared_Value",

    "Redelivery_Flag",

    "SameDay_Flag",

    "StorePickup_Flag",

    "Pickup_Density",

    "Customer_Cold_Ratio",

    "Customer_Remote_Ratio",

    "Customer_Redelivery_Rate"
]


categorical_features = [

    "Package_Size",

    "Temperature",

    "Destination_Region",

    "Customer_Type",

    "Industry"
]


required_columns = (

    numeric_features

    + categorical_features

    + [

        "Actual_Operational_Cost",

        "Customer_ID"
    ]
)


missing_columns = [

    col

    for col in required_columns

    if col not in df.columns
]


if len(missing_columns) > 0:

    raise ValueError(

        "Excel 缺少以下必要欄位："

        + str(missing_columns)
    )


X = df[

    numeric_features

    + categorical_features

].copy()


y = df[
    "Actual_Operational_Cost"
].copy()


# ============================================================
# STEP 5 - Preprocessing
# ============================================================

section_title(

    5,

    "資料前處理",

    """
    使用與 Multiple Regression 相同的資料處理方法，
    數值欄位處理缺值，
    類別欄位進行 One-Hot Encoding。
    """
)


numeric_transformer = Pipeline(

    steps=[

        (

            "imputer",

            SimpleImputer(
                strategy="median"
            )
        )
    ]
)


categorical_transformer = Pipeline(

    steps=[

        (

            "imputer",

            SimpleImputer(
                strategy="most_frequent"
            )
        ),

        (

            "onehot",

            OneHotEncoder(

                handle_unknown="ignore",

                drop="first"
            )
        )
    ]
)


preprocessor = ColumnTransformer(

    transformers=[

        (

            "num",

            numeric_transformer,

            numeric_features
        ),

        (

            "cat",

            categorical_transformer,

            categorical_features
        )
    ]
)


# ============================================================
# STEP 6 - Train / Test Split
# ============================================================

section_title(

    6,

    "依 Customer ID 切分 Training / Testing",

    """
    使用與 Multiple Regression 相同的 random_state=42，
    讓兩種模型面對同一組 Testing Customers，
    確保模型比較公平。
    """
)


groups = df[
    "Customer_ID"
]


splitter = GroupShuffleSplit(

    n_splits=1,

    test_size=0.25,

    random_state=42
)


train_index, test_index = next(

    splitter.split(

        X,

        y,

        groups
    )
)


X_train = X.iloc[
    train_index
].copy()


X_test = X.iloc[
    test_index
].copy()


y_train = y.iloc[
    train_index
].copy()


y_test = y.iloc[
    test_index
].copy()


train_customer_list = sorted(

    df.iloc[
        train_index
    ]["Customer_ID"].unique()
)


test_customer_list = sorted(

    df.iloc[
        test_index
    ]["Customer_ID"].unique()
)


metric_cards([

    (
        "Training Data",

        f"{len(X_train):,}",

        "模型學習資料"
    ),

    (
        "Testing Data",

        f"{len(X_test):,}",

        "模型測試資料"
    ),

    (
        "Training Customers",

        f"{len(train_customer_list)}",

        "訓練客戶"
    ),

    (
        "Testing Customers",

        f"{len(test_customer_list)}",

        "未見新客戶"
    )
])


# ============================================================
# STEP 7 - Build Random Forest
# ============================================================

section_title(

    7,

    "建立 Random Forest 模型",

    """
    Random Forest 建立大量 Decision Trees，
    每一棵樹從不同資料與特徵進行學習，
    最後將多棵樹的預測結果平均，
    形成 Cost-to-Serve 預測。
    """
)


model_rf = Pipeline(

    steps=[

        (

            "preprocessor",

            preprocessor
        ),

        (

            "random_forest",

            RandomForestRegressor(

                n_estimators=300,

                max_depth=None,

                min_samples_split=2,

                min_samples_leaf=1,

                max_features=1.0,

                random_state=42,

                n_jobs=-1
            )
        )
    ]
)


model_rf.fit(

    X_train,

    y_train
)


info_box(

    """
    Random Forest 訓練完成。
    本模型使用 300 棵 Decision Trees，
    並將多棵樹的預測結果平均，
    以降低單一決策樹容易過度擬合的風險。
    """
)


# ============================================================
# STEP 8 - Prediction
# ============================================================

pred_rf = model_rf.predict(
    X_test
)


# ============================================================
# STEP 9 - Model Evaluation
# ============================================================

r2 = r2_score(
    y_test,
    pred_rf
)


mae = mean_absolute_error(
    y_test,
    pred_rf
)


rmse = np.sqrt(

    mean_squared_error(

        y_test,

        pred_rf
    )
)


mape = (

    mean_absolute_percentage_error(

        y_test,

        pred_rf
    )

    * 100
)


section_title(

    8,

    "Random Forest 模型績效",

    """
    使用與 Multiple Regression 相同的
    R²、MAE、RMSE、MAPE 進行評估。
    """
)


metric_cards([

    (

        "R²",

        f"{r2:.4f}",

        "越接近 1 越好"
    ),

    (

        "MAE",

        f"${mae:.2f}",

        "平均成本預測誤差"
    ),

    (

        "RMSE",

        f"${rmse:.2f}",

        "大型預測誤差"
    ),

    (

        "MAPE",

        f"{mape:.2f}%",

        "平均百分比誤差"
    )
])


# ============================================================
# STEP 9 - Package Prediction Results
# ============================================================

result_rf = df.iloc[

    test_index

][
    [

        "Package_ID",

        "Customer_ID",

        "Industry",

        "Customer_Type",

        "Monthly_Target_Packages",

        "Package_Size",

        "Weight_kg",

        "Temperature",

        "Destination_Region",

        "Remote_Flag",

        "Distance_km",

        "COD_Flag",

        "COD_Amount",

        "Redelivery_Flag",

        "SameDay_Flag",

        "Pickup_Density",

        "Actual_Operational_Cost"
    ]
].copy()


result_rf[
    "RF_Predicted_Cost"
] = pred_rf


result_rf[
    "Prediction_Error"
] = (

    result_rf[
        "RF_Predicted_Cost"
    ]

    -

    result_rf[
        "Actual_Operational_Cost"
    ]
)


result_rf[
    "Absolute_Error"
] = abs(

    result_rf[
        "Prediction_Error"
    ]
)


result_rf[
    "Percentage_Error"
] = (

    result_rf[
        "Absolute_Error"
    ]

    /

    result_rf[
        "Actual_Operational_Cost"
    ]

    * 100
)


section_title(

    9,

    "Random Forest 成本預測結果",

    """
    逐筆比較每件包裹的 Actual Cost
    與 Random Forest Predicted Cost。
    """
)


display(

    warm_style(

        result_rf.head(20),

        formats={

            "Weight_kg":
                "{:.2f}",

            "Distance_km":
                "{:.1f}",

            "Actual_Operational_Cost":
                "${:.2f}",

            "RF_Predicted_Cost":
                "${:.2f}",

            "Prediction_Error":
                "${:.2f}",

            "Absolute_Error":
                "${:.2f}",

            "Percentage_Error":
                "{:.2f}%"
        },

        gradient_cols=[

            "Absolute_Error",

            "Percentage_Error"
        ]
    )
)


# ============================================================
# Chart 2 - Actual vs Predicted
# ============================================================

section_title(

    10,

    "Actual Cost vs Predicted Cost",

    """
    不同 Package Size 使用不同顏色。
    點越接近 Perfect Prediction Line，
    代表成本預測越準確。
    """
)


fig, ax = plt.subplots(
    figsize=(9, 7)
)


for size in [

    "S60",

    "S90",

    "S120",

    "S150"
]:

    temp = result_rf[

        result_rf[
            "Package_Size"
        ] == size
    ]


    ax.scatter(

        temp[
            "Actual_Operational_Cost"
        ],

        temp[
            "RF_Predicted_Cost"
        ],

        label=size,

        alpha=0.65,

        s=45,

        color=SIZE_COLORS[size],

        edgecolor="white",

        linewidth=0.5
    )


min_value = min(

    result_rf[
        "Actual_Operational_Cost"
    ].min(),

    result_rf[
        "RF_Predicted_Cost"
    ].min()
)


max_value = max(

    result_rf[
        "Actual_Operational_Cost"
    ].max(),

    result_rf[
        "RF_Predicted_Cost"
    ].max()
)


ax.plot(

    [
        min_value,
        max_value
    ],

    [
        min_value,
        max_value
    ],

    color=COLORS["brown"],

    linestyle="--",

    linewidth=2,

    label="Perfect Prediction"
)


ax.set_title(

    "Random Forest\nActual Cost vs Predicted Cost",

    color=COLORS["brown_dark"],

    pad=16
)


ax.set_xlabel(
    "Actual Operational Cost"
)


ax.set_ylabel(
    "Random Forest Predicted Cost"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    alpha=0.15
)


ax.legend(
    frameon=False,
    title="Package Size"
)


plt.tight_layout()

plt.show()


# ============================================================
# STEP 11 - Package Size Error Analysis
# ============================================================

size_error = (

    result_rf

    .groupby(
        "Package_Size"
    )

    .agg(

        Package_Count=(

            "Package_ID",

            "count"
        ),

        Actual_Avg_Cost=(

            "Actual_Operational_Cost",

            "mean"
        ),

        Predicted_Avg_Cost=(

            "RF_Predicted_Cost",

            "mean"
        ),

        MAE=(

            "Absolute_Error",

            "mean"
        ),

        MAPE=(

            "Percentage_Error",

            "mean"
        )
    )

    .reindex([

        "S60",

        "S90",

        "S120",

        "S150"
    ])

    .reset_index()

    .round(2)
)


section_title(

    11,

    "不同 Package Size 的成本預測誤差",

    """
    比較 Random Forest 在
    S60、S90、S120、S150
    不同材積級距上的預測能力。
    """
)


display(

    warm_style(

        size_error,

        formats={

            "Package_Count":
                "{:,.0f}",

            "Actual_Avg_Cost":
                "${:.2f}",

            "Predicted_Avg_Cost":
                "${:.2f}",

            "MAE":
                "${:.2f}",

            "MAPE":
                "{:.2f}%"
        },

        gradient_cols=[
            "MAE",
            "MAPE"
        ]
    )
)


# ============================================================
# Chart 3 - Package Size Error
# ============================================================

fig, ax = plt.subplots(
    figsize=(9, 5.5)
)


bars = ax.bar(

    size_error[
        "Package_Size"
    ],

    size_error[
        "MAE"
    ],

    color=[

        SIZE_COLORS[x]

        for x in

        size_error[
            "Package_Size"
        ]
    ],

    edgecolor="white"
)


ax.set_title(
    "Random Forest Prediction Error by Package Size",
    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Package Size"
)


ax.set_ylabel(
    "Mean Absolute Error"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


for bar, value in zip(
    bars,
    size_error["MAE"]
):

    ax.text(

        bar.get_x()
        + bar.get_width()/2,

        value + 0.1,

        f"${value:.2f}",

        ha="center",

        fontweight="bold",

        color=COLORS["brown"]
    )


plt.tight_layout()

plt.show()


# ============================================================
# STEP 12 - Temperature Analysis
# ============================================================

temperature_error = (

    result_rf

    .groupby(
        "Temperature"
    )

    .agg(

        Package_Count=(

            "Package_ID",

            "count"
        ),

        Actual_Avg_Cost=(

            "Actual_Operational_Cost",

            "mean"
        ),

        Predicted_Avg_Cost=(

            "RF_Predicted_Cost",

            "mean"
        ),

        MAE=(

            "Absolute_Error",

            "mean"
        )
    )

    .reset_index()

    .round(2)
)


section_title(

    12,

    "常溫與低溫成本預測比較",

    """
    分析 Random Forest
    對 Ambient 與 Chilled
    不同物流情境的成本預測能力。
    """
)


display(

    warm_style(

        temperature_error,

        formats={

            "Package_Count":
                "{:,.0f}",

            "Actual_Avg_Cost":
                "${:.2f}",

            "Predicted_Avg_Cost":
                "${:.2f}",

            "MAE":
                "${:.2f}"
        },

        gradient_cols=[
            "MAE"
        ]
    )
)


temperature_plot = (
    temperature_error.copy()
)


temperature_plot[
    "Temperature_EN"
] = (

    temperature_plot[
        "Temperature"
    ]

    .map(
        TEMPERATURE_EN
    )

    .fillna(
        temperature_plot[
            "Temperature"
        ]
    )
)


# ============================================================
# Chart 4 - Temperature
# ============================================================

fig, ax = plt.subplots(
    figsize=(7.5, 5)
)


bars = ax.bar(

    temperature_plot[
        "Temperature_EN"
    ],

    temperature_plot[
        "MAE"
    ],

    color=[

        TEMP_COLORS.get(
            x,
            COLORS["peach"]
        )

        for x in

        temperature_plot[
            "Temperature_EN"
        ]
    ],

    edgecolor="white"
)


ax.set_title(
    "Random Forest Prediction Error by Temperature",
    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Temperature Type"
)


ax.set_ylabel(
    "Mean Absolute Error"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


for bar, value in zip(
    bars,
    temperature_plot["MAE"]
):

    ax.text(

        bar.get_x()
        + bar.get_width()/2,

        value + 0.1,

        f"${value:.2f}",

        ha="center",

        fontweight="bold",

        color=COLORS["brown"]
    )


plt.tight_layout()

plt.show()


# ============================================================
# STEP 13 - Region Analysis
# ============================================================

region_error = (

    result_rf

    .groupby(
        "Destination_Region"
    )

    .agg(

        Package_Count=(

            "Package_ID",

            "count"
        ),

        Actual_Avg_Cost=(

            "Actual_Operational_Cost",

            "mean"
        ),

        Predicted_Avg_Cost=(

            "RF_Predicted_Cost",

            "mean"
        ),

        MAE=(

            "Absolute_Error",

            "mean"
        )
    )

    .reset_index()

    .round(2)
)


section_title(

    13,

    "不同配送區域成本預測比較",

    """
    比較 North、Central、South、
    East 與 Offshore 的模型誤差。
    """
)


display(

    warm_style(

        region_error,

        formats={

            "Package_Count":
                "{:,.0f}",

            "Actual_Avg_Cost":
                "${:.2f}",

            "Predicted_Avg_Cost":
                "${:.2f}",

            "MAE":
                "${:.2f}"
        },

        gradient_cols=[
            "MAE"
        ]
    )
)


region_plot = (
    region_error.copy()
)


region_plot[
    "Region_EN"
] = (

    region_plot[
        "Destination_Region"
    ]

    .map(
        REGION_EN
    )

    .fillna(
        region_plot[
            "Destination_Region"
        ]
    )
)


region_order = [

    "North",

    "Central",

    "South",

    "East",

    "Offshore"
]


region_plot[
    "Region_EN"
] = pd.Categorical(

    region_plot[
        "Region_EN"
    ],

    categories=region_order,

    ordered=True
)


region_plot = (

    region_plot

    .sort_values(
        "Region_EN"
    )
)


# ============================================================
# Chart 5 - Region Error
# ============================================================

fig, ax = plt.subplots(
    figsize=(9, 5.5)
)


bars = ax.bar(

    region_plot[
        "Region_EN"
    ].astype(str),

    region_plot[
        "MAE"
    ],

    color=[

        REGION_COLORS.get(
            str(x),
            COLORS["peach"]
        )

        for x in

        region_plot[
            "Region_EN"
        ]
    ],

    edgecolor="white"
)


ax.set_title(
    "Random Forest Prediction Error by Destination Region",
    color=COLORS["brown_dark"]
)


ax.set_xlabel(
    "Destination Region"
)


ax.set_ylabel(
    "Mean Absolute Error"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="y",
    alpha=0.15
)


for bar, value in zip(
    bars,
    region_plot["MAE"]
):

    ax.text(

        bar.get_x()
        + bar.get_width()/2,

        value + 0.1,

        f"${value:.2f}",

        ha="center",

        fontweight="bold",

        color=COLORS["brown"]
    )


plt.tight_layout()

plt.show()


# ============================================================
# STEP 14 - Worst Cases
# ============================================================

worst_cases = (

    result_rf

    .sort_values(

        "Absolute_Error",

        ascending=False
    )

    .head(20)
)


section_title(

    14,

    "Random Forest 預測誤差最大的 20 件包裹",

    """
    找出模型最難處理的物流情境，
    並與 Multiple Regression 的 Worst Cases 比較。
    """
)


display(

    warm_style(

        worst_cases,

        formats={

            "Weight_kg":
                "{:.2f}",

            "Distance_km":
                "{:.1f}",

            "Actual_Operational_Cost":
                "${:.2f}",

            "RF_Predicted_Cost":
                "${:.2f}",

            "Prediction_Error":
                "${:.2f}",

            "Absolute_Error":
                "${:.2f}",

            "Percentage_Error":
                "{:.2f}%"
        },

        gradient_cols=[

            "Absolute_Error",

            "Percentage_Error"
        ]
    )
)


# ============================================================
# STEP 15 - Feature Importance
# ============================================================

section_title(

    15,

    "Random Forest Feature Importance",

    """
    與 Multiple Regression 的 Regression Coefficient 不同，
    Random Forest 使用 Feature Importance
    觀察哪些物流特徵對模型判斷最重要。
    """
)


fitted_preprocessor = (

    model_rf

    .named_steps[
        "preprocessor"
    ]
)


feature_names = (

    fitted_preprocessor

    .get_feature_names_out()
)


feature_names = [

    str(x)

    .replace(
        "num__",
        ""
    )

    .replace(
        "cat__",
        ""
    )

    for x in feature_names
]


importance_values = (

    model_rf

    .named_steps[
        "random_forest"
    ]

    .feature_importances_
)


importance_df = pd.DataFrame({

    "Feature":
        feature_names,

    "Importance":
        importance_values
})


importance_df[
    "Feature_EN"
] = (

    importance_df[
        "Feature"
    ]

    .apply(
        feature_name_to_english
    )
)


importance_df = (

    importance_df

    .sort_values(

        "Importance",

        ascending=False
    )
)


top_importance = (

    importance_df

    .head(20)

    .copy()
)


display(

    warm_style(

        top_importance[
            [
                "Feature",
                "Importance"
            ]
        ],

        formats={

            "Importance":
                "{:.4f}"
        },

        gradient_cols=[
            "Importance"
        ]
    )
)


# ============================================================
# Chart 6 - Feature Importance
# ============================================================

importance_chart = (

    top_importance

    .sort_values(
        "Importance"
    )
)


fig, ax = plt.subplots(
    figsize=(11, 8)
)


ax.barh(

    importance_chart[
        "Feature_EN"
    ],

    importance_chart[
        "Importance"
    ],

    color=COLORS["coral"]
)


ax.set_title(

    "Top Cost Drivers\nRandom Forest Feature Importance",

    color=COLORS["brown_dark"],

    pad=15
)


ax.set_xlabel(
    "Feature Importance"
)


ax.set_ylabel(
    "Cost Driver"
)


ax.set_facecolor(
    COLORS["cream"]
)


ax.grid(
    axis="x",
    alpha=0.15
)


plt.tight_layout()

plt.show()


# ============================================================
# STEP 16 - Model Summary
# ============================================================

model_summary = pd.DataFrame({

    "Model": [
        "Random Forest"
    ],

    "R2": [
        r2
    ],

    "MAE_Cost": [
        mae
    ],

    "RMSE_Cost": [
        rmse
    ],

    "MAPE_Cost_%": [
        mape
    ],

    "Trees": [
        300
    ],

    "Training_Records": [
        len(X_train)
    ],

    "Testing_Records": [
        len(X_test)
    ],

    "Training_Customers": [
        len(train_customer_list)
    ],

    "Testing_Customers": [
        len(test_customer_list)
    ]
})


section_title(

    16,

    "Random Forest 模型總結",

    """
    完成 Random Forest Cost-to-Serve 預測，
    後續將使用完全相同的方法建立 XGBoost，
    最後進行三模型公平比較。
    """
)


display(

    warm_style(

        model_summary,

        formats={

            "R2":
                "{:.4f}",

            "MAE_Cost":
                "${:.2f}",

            "RMSE_Cost":
                "${:.2f}",

            "MAPE_Cost_%":
                "{:.2f}%"
        },

        gradient_cols=[
            "R2"
        ]
    )
)


# ============================================================
# STEP 17 - Export Excel
# ============================================================

output_filename = (
    "Logistics_WMS_Random_Forest_Cost_Prediction.xlsx"
)


with pd.ExcelWriter(

    output_filename,

    engine="openpyxl"

) as writer:


    model_summary.to_excel(

        writer,

        sheet_name="Model_Summary",

        index=False
    )


    result_rf.to_excel(

        writer,

        sheet_name="Package_Predictions",

        index=False
    )


    worst_cases.to_excel(

        writer,

        sheet_name="Worst_Cases",

        index=False
    )


    importance_df.to_excel(

        writer,

        sheet_name="Feature_Importance",

        index=False
    )


    size_error.to_excel(

        writer,

        sheet_name="Size_Analysis",

        index=False
    )


    temperature_error.to_excel(

        writer,

        sheet_name="Temperature_Analysis",

        index=False
    )


    region_error.to_excel(

        writer,

        sheet_name="Region_Analysis",

        index=False
    )


# ============================================================
# STEP 18 - Excel Styling
# ============================================================

from openpyxl import load_workbook

from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)


wb = load_workbook(
    output_filename
)


header_fill = PatternFill(

    "solid",

    fgColor="C96F5B"
)


alternate_fill = PatternFill(

    "solid",

    fgColor="FFF8F0"
)


header_font = Font(

    color="FFFFFF",

    bold=True
)


body_font = Font(

    color="4F3A30"
)


thin_border = Border(

    bottom=Side(

        style="thin",

        color="EBDDD3"
    )
)


for ws in wb.worksheets:


    ws.freeze_panes = "A2"


    ws.auto_filter.ref = ws.dimensions


    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(

            horizontal="center",

            vertical="center"
        )


    for row in range(
        2,
        ws.max_row + 1
    ):


        if row % 2 == 0:

            for cell in ws[row]:

                cell.fill = alternate_fill


        for cell in ws[row]:

            cell.font = body_font

            cell.border = thin_border

            cell.alignment = Alignment(
                vertical="center"
            )


    for column_cells in ws.columns:


        length = 0


        column_letter = (
            column_cells[0]
            .column_letter
        )


        for cell in column_cells:

            try:

                value = str(

                    cell.value

                    if cell.value is not None

                    else ""
                )


                length = max(
                    length,
                    len(value)
                )

            except:

                pass


        ws.column_dimensions[
            column_letter
        ].width = min(

            max(
                length + 3,
                12
            ),

            32
        )


    ws.row_dimensions[1].height = 24


wb.save(
    output_filename
)


# ============================================================
# Final Dashboard
# ============================================================

dashboard_title(

    "物流業 WMS｜Random Forest 分析完成",

    """
    已完成 Random Forest Cost-to-Serve 成本預測，
    包含模型 KPI、Package Size、Temperature、
    Destination Region、Worst Cases
    與 Feature Importance 分析。
    """
)


metric_cards([

    (
        "R²",

        f"{r2:.4f}",

        "模型解釋能力"
    ),

    (
        "MAE",

        f"${mae:.2f}",

        "平均成本誤差"
    ),

    (
        "RMSE",

        f"${rmse:.2f}",

        "大型誤差風險"
    ),

    (
        "MAPE",

        f"{mape:.2f}%",

        "平均百分比誤差"
    )
])


info_box(

    """
    下一階段將建立 XGBoost。
    XGBoost 與 Random Forest 都以 Decision Tree 為基礎，
    但 Random Forest 是「多棵樹平行獨立學習後平均」，
    XGBoost 則是「後面的樹持續修正前面模型的預測誤差」。
    """
)


print(
    "📁 Analysis file:",
    output_filename
)


files.download(
    output_filename
)
